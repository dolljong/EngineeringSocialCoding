#import xlwings as xw
import openpyxl as op
from openpyxl import load_workbook, Workbook
from string import ascii_uppercase
from openpyxl.styles.fonts import Font
from openpyxl.styles import Border, Side, Alignment, PatternFill
from openpyxl.styles.colors import Color
from sec_back_ver031 import *
from tkinter import filedialog

import xlwings as xw



def readxls():
    wb = xw.books.active
    #sheet = xw.sheets.active
    #print(wb.name)
    #print(sheet.name)    
    
    wsin = xw.sheets.active
    #wsin = wb['Sheet1']
    data_rc_mat = []  #재료특성 입력 data_rc_mat
    data_mat_fac_uls = [] #재료계수(극한한계상태) 입력 data_matfac_uls 
    data_mat_fac_als = [] #재료계수(극단상황한계상태) 입력 data_matfac_als
    data_sec = [] #단면제원 입력 data_sec
    data_usedas_band = [] #사용휨철근량 입력 data_usedas_band
    data_usedas_shear = [] #전단철근량 및 각도 입력 data_usedas_shear
    data_secforce = [] #산정하중(극한한계상태) 입력 data_secforce_uls
    #data_secforce_als = [] #산정하중(극단상황한계상태) 입력 data_secforce_als

    data_rc_mat = wsin['C4:D4'].value       #재료특성 입력

    data_mat_fac_uls = wsin['C7:D7'].value  #재료계수(극한한계상태) 입력

    data_mat_fac_als = wsin['C8:D8'].value  #재료계수(극단상황한계상태) 입력

    print(data_rc_mat,data_mat_fac_uls,data_mat_fac_als)

    expdata = wsin['C12'].expand('table').value
    print(expdata)

    lastsheetname="Sheet1"
    for row in expdata:
        data_case_name = row[0]
        data_sec = row[1:3]  #단면제원 입력
        data_secforce = row[3:10]
        #data_secforce_als = row[:]
        #data_ulsorals = row[9]
        data_usedas_band = row[10:19]
        data_usedas_shear = row[19:25]

        calc = Sec_back(data_rc_mat, data_mat_fac_uls, data_sec, data_usedas_band, data_usedas_shear, data_secforce)
        wsout = wb.sheets.add(data_case_name, after= lastsheetname)
        wsout.activate
        
        wsinit(wsout)

        calc.calmoment()
        momentexcelout(calc, wsout)
        
        calc.calshear()
        shearexcelout(calc, wsout)

        calc.calservice()
        serviceexcelout(calc, wsout)

        lastsheetname = data_case_name


        #print(data_case_name,data_sec,data_secforce,data_usedas_band,data_usedas_shear)

    #return
    
    # datatp9 = wsin['C11:D11']                 #단면제원 입력
    # for datatp10 in datatp9 :                 #튜플형을 리스트형으로 변환 
    #     for datatp11 in datatp10 :
    #         data_sec.append(datatp11.value)
    # datatp12 = wsin['C15:K15']                 #사용휨철근량 입력  
    # for datatp13 in datatp12 :                 #튜플형을 리스트형으로 변환 
    #     for datatp14 in datatp13 :
    #         data_usedas_band.append(datatp14.value)
    # datatp15 = wsin['C18:H18']                 #전단철근량 및 각도 입력
    # for datatp16 in datatp15 :                 #튜플형을 리스트형으로 변환 
    #     for datatp17 in datatp16 :
    #         data_usedas_shear.append(datatp17.value)
    # datatp18 = wsin['C21:H21']                 #산정하중(극한한계상태) 입력
    # for datatp19 in datatp18 :                 #튜플형을 리스트형으로 변환 
    #     for datatp20 in datatp19 :
    #         data_secforce_uls.append(datatp20.value)
    # datatp21 = wsin['C22:H22']                 #산정하중(극단상황한계상태) 입력
    # for datatp22 in datatp21 :                 #튜플형을 리스트형으로 변환 
    #     for datatp23 in datatp22 :
    #         data_secforce_als.append(datatp23.value)
    #calc = Sec_back(data_rc_mat, data_mat_fac_uls, data_sec, data_usedas_band, data_usedas_shear, data_secforce)
    #calc1 = Sec_back(data_rc_mat, data_mat_fac_als, data_sec, data_usedas_band, data_usedas_shear, data_secforce)

    #wb.close()

    #return #xlsdata    

def wsinit(wsout) :                             #워크시트 형식 초기화
    alpalist = list(ascii_uppercase)
    #for i in range(1,100) :                     # 헹 크기 지정
        #wsout.row_dimensions[i].height = 15
    wsout.range("A1:A200").row_height = 15
    #for i in alpalist :                         # 열 크기 지정
        #wsout.column_dimensions[i].width = 3.0
    wsout.range("A1:Z1").column_width = 3.0
    font_format = Font(size=9, name = '굴림체')
    # for rows in wsout["A1":"Z200"] :            # 기본 폰트를 size=9, 굴림체로 변경
    #     for cell in rows :
    #         cell.font = font_format
    wsout.range("A1:Z200").font.name = "굴림체"
    wsout.range("A1:Z200").font.size = 9
    # wsout.range["A1:Z100"].font.name = "굴림체"
    # wsout.range["A1:Z100"].font.size = 9

def momentexcelout(calc, wsout) :
    wsout['B2'].value = '1) 단면제원 및 설계가정'
    wsout['C3'].value ="fck = %d Mpa, fy = %d Mpa, Øc = %1.2f, Øs = %1.2f, Es = %d Mpa" %(calc.fck, calc.fy, calc.Øc, calc.Øs, calc.Es)

    for r in range(4,6) :                        # 셀 병합 C4~L4, C5~L5
        for c in [3,6,9,12] :                       
            c1 = c + 2
            #wsout.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c1)
            wsout.range((r,c), (r,c1)).merge()

    for r in range(4,6) :                         # 셀 병합 O4~T4, O5~T5  
        for c in [15,20] :                          
            c1 = c + 4
            #wsout.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c1)  
            wsout.range((r,c), (r, c1)).merge()   

    # for r in range(4,6) :                        # 4~5행 글짜 위치 중앙으로 정렬
    #     for c in range(3,24) :                      
    #         #wsout.cell(r,c).alignment = Alignment(horizontal='center', vertical='center')
    #         wsout.range(r,c).api.HorizontalAlignment = xw.constants.Constants.xlCenter 
    wsout.range("C4:X5").api.HorizontalAlignment = xw.constants.Constants.xlCenter 
    
    # for r in range(4,6) :                        # 4~5행 테두리 그리기
    #     for c in range(3,24) :                      
    #         wsout.range(r,c).border = Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    bordertop = xw.constants.BordersIndex.xlEdgeTop
    borderbottom = xw.constants.BordersIndex.xlEdgeBottom
    borderleft = xw.constants.BordersIndex.xlEdgeLeft
    borderright = xw.constants.BordersIndex.xlEdgeRight
    borderinsideH = xw.constants.BordersIndex.xlInsideHorizontal
    borderinsideV = xw.constants.BordersIndex.xlInsideVertical
    wsout.range("C4:X5").api.Borders(bordertop).Weight = 4
    wsout.range("C4:X5").api.Borders(borderbottom).Weight = 4
    wsout.range("C4:X5").api.Borders(borderleft).Weight = 4
    wsout.range("C4:X5").api.Borders(borderright).Weight = 4 
    wsout.range("C4:X5").api.Borders(borderinsideH).Weight = 2 
    wsout.range("C4:X5").api.Borders(borderinsideV).Weight = 2
 
    # for c in range(3,24) :                         # 셀 채우기 C4~T5
    #     wsout.range(4,c).fill = PatternFill(fill_type='solid', fgColor='0000FFFF')
    wsout.range("C4:X4").color = (0,255,255)

    wsout['C4'].value = 'B(mm)'
    wsout['F4'].value = 'H(mm)'
    wsout['I4'].value = 'd(mm)'
    wsout['L4'].value = '피복(mm)'
    wsout['O4'].value = 'Mu(N.mm)'
    wsout['T4'].value = 'Vu(N)'
    wsout['C5'].value = calc.B
    wsout['F5'].value = calc.H
    wsout['I5'].value = "%4.3f" %(calc.D)
    wsout['L5'].value = "%4.3f" %(calc.Dc)
    wsout['O5'].value = calc.Mun
    wsout['O5'].number_format = '#,##0#'
    wsout['T5'].value = calc.Vun
    wsout['T5'].number_format = '#,##0#'
    wsout['B7'].value = '2) 콘크리트 재료상수'
    for r in range(8,19) :
        #wsout.merge_cells(start_row=r, start_column=17, end_row=r, end_column=18)
        wsout.range((r, 17), (r, 18)).merge()
    wsout['C8'].value = 'n      : 상승 곡선부의 형상을 나타내는 지수'
    wsout['P8'].value = '='
    wsout['Q8'].value = calc.nε
    wsout['C9'].value = 'εco,r : 최대응력에 처음 도달했을때의 변형률'
    wsout['P9'].value = '='
    wsout['Q9'].value = calc.εco
    wsout['C10'].value = 'εcu,r : 극한변형률'
    wsout['P10'].value = '='
    wsout['Q10'].value = calc.εcu
    wsout['C11'].value = 'αcc   : 유효계수'
    wsout['P11'].value = '='
    wsout['Q11'].value = calc.αcc
    wsout['C12'].value = 'fcd   : 콘크리트 설계압축강도'
    wsout['P12'].value = '='
    wsout['Q12'].value = calc.fcd
    wsout['S12'].value = 'MPa'
    wsout['C13'].value = 'fcm   : 평균압축강도(fck+Δf)'
    wsout['P13'].value = '='
    wsout['Q13'].value = calc.fcm
    wsout['S13'].value = 'MPa'
    wsout['C14'].value = 'Ec    : 콘크리트 탄성계수'
    wsout['P14'].value = '='
    wsout['Q14'].value = calc.Ec
    wsout['S14'].value = 'MPa'
    wsout['C15'].value = 'α     : 압축합력의 평균 응력계수'
    wsout['P15'].value = '='
    wsout['Q15'].value = calc.α
    wsout['C16'].value = 'β     : 압축합력의 작용점 위치계수'
    wsout['P16'].value = '='
    wsout['Q16'].value = calc.β
    wsout['C17'].value = 'η     : 등가 사각형 응력 블록의 크기계수'
    wsout['P17'].value = '='
    wsout['Q17'].value = calc.η
    wsout['C18'].value = 'β1    : 등가 사각형 응력 블록의 깊이계수(2β)'
    wsout['P18'].value = '='
    wsout['Q18'].value = calc.β1

    wsout['B20'].value = '3) 철근 재료상수'
    for r in range(21,23) :
        wsout.range((r,17), (r, 18)).merge()
    wsout['C21'].value = 'fyd    : 설계인장강도 ( Φs fy )'
    wsout['P21'].value = '='
    wsout['Q21'].value = calc.fyd
    wsout['S21'].value = 'MPa'
    wsout['C22'].value = 'εyd    : 설계 항복 변형률 ( fyd / Es )'
    wsout['P22'].value = '='
    wsout['Q22'].value = calc.εyd

    wsout['B24'].value = '4) 필요철근량 산정'
    wsout['C25'].value = 'Mu = As x fyd  x (d - a / 2)              ----------------   ①'
    wsout['C26'].value = ' a = As x fyd  / ( η x fcd x b)          ----------------   ②'
    wsout['C27'].value = ' 식②를 식①에 대입하여 이차방정식으로 As를 구한다'
    wsout['F28'].value = ' fyd²'
    wsout['C29'].value = ' ──────────── As² - fyd x d x As + Mu = 0 ,   Asreq = %6.1fmm²' %(calc.Asreq)
    wsout['D30'].value = ' 2 x η x fcd x b'

    wsout['B32'].value ="5) 사용철근량 : Asuse =  %6.1f mm², 철근도심 : dc =  %5.1f mm [ 사용율 = %3.3f ]" %(calc.Asuse, calc.Dc, calc.Asuse/calc.Asreq)
    wsout['F33'].value = "1단 : %c %d - %d EA (=  %6.1f mm², dc1 =  %4.1f mm)" %(calc.rebarid, calc.AsDia1, calc.AsNum1, calc.Asuse1*calc.AsNum1, calc.Dc1)
    wsout['F34'].value = "2단 : %c %d - %d EA (=  %6.1f mm², dc2 =  %4.1f mm)" %(calc.rebarid, calc.AsDia2, calc.AsNum2, calc.Asuse2*calc.AsNum2, calc.Dc2)
    wsout['F35'].value = "3단 : %c %d - %d EA (=  %6.1f mm², dc3 =  %4.1f mm)" %(calc.rebarid, calc.AsDia3, calc.AsNum3, calc.Asuse3*calc.AsNum3, calc.Dc3)

    wsout['B37'].value ="6) 철근량 검토"
    wsout['C38'].value ="As,min = (0.25 √fck / fy) x b x d = %6.1f mm²" %(calc.Asmin1)
    wsout['D39'].value ="   = (1.4 / fy) x b x d = %6.1f mm²" %(calc.Asmin2) 
    wsout['D40'].value ="   = As_req x 4 / 3 = %6.1f mm²" %(calc.Asmin3) 
    if calc.Asuse >= calc.Asmin :
        wsout['C41'].value ="As,use = %6.1f mm² ≥ As,min = %6.1f mm²  ∴ O.K" %(calc.Asuse, calc.Asmin)
    else :
        wsout['C41'].value ="As,use = %6.1f mm² < As,min = %6.1f mm²  ∴ N.G" %(calc.Asuse, calc.Asmin)
    if calc.Asuse <= calc.Asmax :
        wsout['C42'].value ="As,max = 0.04 x b x d = %6.1f mm² ≥  As,use = %6.1f mm²  ∴ O.K" %(calc.Asmax, calc.Asuse)
    else :
        wsout['C42'].value ="As,max = 0.04 x b x d = %6.1f mm² <  As,use = %6.1f mm²  ∴ N.G" %(calc.Asmax, calc.Asuse)

    wsout['B44'].value ="7) 중립축 깊이 검토"
    wsout['C45'].value ="Cmax = (δ x εcu / 0.0033 - 0.6) x d" 
    wsout['D46'].value =" = (%2.1f x %2.5f / 0.0033 - 0.6) x %4.1f = %4.1f mm" %(calc.δ, calc.εcu, calc.D, calc.c_max)
    wsout['C47'].value ="C = Φs x As x fy / (α x Φc x 0.85 x fck x b)" 
    wsout['C48'].value ="  = %2.2f x %6.1f x %d / (%2.2f x %2.2f x 0.85 x %d x %5.1f)" %(calc.Øs, calc.Asuse, calc.fy, calc.α, calc.Øc, calc.fck, calc.B)
    if calc.cc <= calc.c_max :
        wsout['C49'].value ="  = %6.1f mm < Cmax = %6.1f mm  ∴ O.K" %(calc.cc, calc.c_max)
    else :
        wsout['C49'].value ="  = %6.1f mm ≥ Cmax = %6.1f mm  ∴ N.G" %(calc.cc, calc.c_max)

    wsout['B51'].value ="8) 인장철근 변형률"
    wsout['C52'].value ="εs = (d - C) / C x εcu" 
    wsout['C53'].value ="    = (%4.1f - %4.1f) / %4.1f x %2.5f = %2.5f " %(calc.D, calc.cc, calc.cc, calc.εcu, calc.εs)
    wsout['C54'].value ="εyd = Φs x fy / Es" 
    if calc.εyd <= calc.εs :
        wsout['D55'].value =" = %2.2f x %d / %d = %2.5f  ≤ εs  ∴ 항복가정 O.K" %(calc.Øs, calc.fy, calc.Es, calc.εyd)
    else :
        wsout['D55'].value =" = %2.2f x %d / %d = %2.5f  > εs  ∴ 항복가정 N.G" %(calc.Øs, calc.fy, calc.Es, calc.εyd)

    wsout['B57'].value ="9) 설계 휨강도 산정"
    wsout['C58'].value ="Mr = As x Φs x fy x (d - β x c)" 
    wsout['C59'].value ="   = %6.1f x %2.2f x %d x (%4.1f - %2.2f x %4.1f)" %(calc.Asuse, calc.Øs, calc.fy, calc.D, calc.β, calc.cc)
    if calc.Mr > calc.Mun :
        wsout['C60'].value ="   = %10.1f N.mm  ≥ Mu = %10.1f N.mm  ∴ O.K  [S.F = %3.3f]" %(calc.Mr, calc.Mun, calc.Msf)
    else :
        wsout['C60'].value ="   = %10.1f N.mm  < Mu = %10.1f N.mm  ∴ N.G  [S.F = %3.3f]" %(calc.Mr, calc.Mun, calc.Msf)


def shearexcelout(calc, wsout) :
    wsout['B62'].value ="10) 전단검토"
    wsout['C63'].value ="Vcd = [0.85 x Φc x k x (p x fck)⅓ + 0.15 x fn] x b x d" 
    wsout['C64'].value ="    = [0.85 x %1.2f x %1.3f x (%1.6f x %3.1f)⅓ + 0.15 x %4.3f] x %4.1f x %4.1f" %(calc.Øc, calc.k, calc.ρs, calc.fck, calc.fn, calc.B, calc.D)
    wsout['C65'].value ="    = %10.0f N" %(calc.Vc) 
    wsout['D66'].value ="k = 1 + √(200 / d) = 1 + √(200 / %4.1f) = %1.3f (≤ 2.000)" %(calc.D, calc.k)
    wsout['D67'].value ="p = As / (b x d) = %6.1f / (%4.1f x %4.1f) = %1.6f ≤ 0.0200" %(calc.Asuse, calc.B, calc.D, calc.ρ) 
    wsout['D68'].value ="fn = Nu / Ac = %4.3f MPa (≤ 0.2 Φc fck = %4.3f MPa, 압축의경우+)" %(calc.fnn, calc.fnmax)
    wsout['C69'].value ="Vcd,min = (0.4 x Φc x fctk + 0.15 x fn] x b x d" 
    wsout['C70'].value ="        = (0.4 x %1.2f x %3.3f + 0.15 x %4.3f] x %4.1f x %4.1f" %(calc.Øc, calc.fctk, calc.fn, calc.B, calc.D)
    wsout['C71'].value ="        = %9.0f N" %(calc.Vcdmin)

    if calc.Vcd >= calc.Vun :
        wsout['C72'].value ="Vcd = %9.0f N ≥ Vu = %9.0f N  ∴ 전단보강 불필요" %(calc.Vcd, calc.Vun)
        wsout['B74'].value ="11) 최소 전단철근량 검토"
        wsout['C75'].value ="Av,use = %5.1f mm² ( %c %d - %2.3fea, C.T.C = %4.1f mm )" %(calc.Avs, calc.rebarid, calc.AvDia, calc.AvLeg, calc.AvSpace)
        wsout['C76'].value ="pv,min = (0.08 x √fck) / fvy = (0.08 x √%3.1f) / %4.1f = %1.6f" %(calc.fck, calc.fy, calc.ρvmin)
        if calc.ρvuse >= calc.ρvmin :
            wsout['C77'].value ="pv,use = %4.1f / (%4.1f x %4.1f x sin%3.1f) = %1.6f ≥ %1.6f ∴ O.K" %(calc.Avs, calc.AvSpace, calc.B, calc.αv, calc.ρvuse, calc.ρvmin)
        else :
            wsout['C77'].value ="pv,use = %4.1f / (%4.1f x %4.1f x sin%3.1f) = %1.6f < %1.6f ∴ N.G" %(calc.Avs, calc.AvSpace, calc.B, calc.αv, calc.ρvuse, calc.ρvmin)
        wsout['B79'].value ="12) 전단철근 간격검토"
        if calc.s1max >= calc.AvSpace :
            wsout['C80'].value ="- 종방향 Smax = 0.75 x %4.1f x (1 + cot%3.1f) = %4.1f ≥ %4.1f   ∴ O.K" %(calc.D, calc.αv, calc.s1max, calc.AvSpace)
        else :
            wsout['C80'].value ="- 종방향 Smax = 0.75 x %4.1f x (1 + cot%3.1f) = %4.1f < %4.1f   ∴ N.G" %(calc.D, calc.αv, calc.s1max, calc.AvSpace)    
        if calc.s2max >= calc.s2 :
            wsout['C81'].value ="- 횡방향 Smax = min(0.75 x %4.1f, 600.0) = %4.1f ≥ %4.1f   ∴ O.K" %(calc.D, calc.s2max, calc.s2)
        else :
            wsout['C81'].value ="- 종방향 Smax = min(0.75 x %4.1f, 600.0) = %4.1f < %4.1f   ∴ N.G" %(calc.D, calc.s2max, calc.s2)
    else :
        wsout['C72'].value ="Vcd = %9.0f N < Vu = %9.0f N  ∴ 전단보강 필요" %(calc.Vcd, calc.Vun)
        wsout['C74'].value ="cotΘ = %1.3f ( Θ = %3.1f˚)" %(calc.cotθ, calc.θ)
        if calc.cotθ < 1 :
            wsout['L74'].value ="1.0 > cotΘ → Θ 불만족  ∴ N.G"
        elif calc.cotθ > 2.5 :    
            wsout['L74'].value ="2.5 ≤ cotΘ → Θ 불만족  ∴ N.G"
        else :
            wsout['L74'].value ="1.0 ≤ cotΘ ≤ 2.5 → Θ 만족  ∴ O.K"
        wsout['C76'].value ="Av,use = %5.1f mm² ( %c %d - %2.3fea, C.T.C = %4.1f mm )" %(calc.Avs, calc.rebarid, calc.AvDia, calc.AvLeg, calc.AvSpace)
        wsout['C77'].value ="Vsd = (Φs x fvy x Av x z / s) x cotθ"
        wsout['C78'].value ="    = (%1.2f x %4.1f x %5.1f x %4.1f / %4.1f) x %1.3f" %(calc.Øs, calc.fy, calc.Avs, calc.z, calc.AvSpace, calc.cotθ)
        if calc.Vd >= calc.Vun :    
            wsout['C79'].value ="    = %d N ≥ Vu = %d N  ∴ O.K" %(calc.Vd, calc.Vun)
        else :
            wsout['C79'].value ="    = %d N < Vu = %d N  ∴ N.G" %(calc.Vd, calc.Vun)    
        wsout['C80'].value ="Vsd,max = (v x Φc x fck x b x z) / (cotθ + tanθ)"
        wsout['C81'].value ="        = (%1.3f x %1.2f x %3.1f x %4.1f x %4.1f) / ( %1.3f + %1.3f )" %(calc.ν, calc.Øc, calc.fck, calc.B, calc.z, calc.cotθ, calc.tanθ)
        if calc.Vd <= calc.Vdmax :
            wsout['C82'].value ="        = %d N ≤ Vsd = %d N  ∴ O.K" %(calc.Vdmax, calc.Vd)
        else :
            wsout['C82'].value ="        = %d N > Vsd = %d N  ∴ N.G" %(calc.Vdmax, calc.Vd)
        wsout['B84'].value ="11) 최소 전단철근량 검토"
        wsout['C85'].value ="pv,min = (0.08 x √fck) / fvy = (0.08 x √%3.1f) / %4.1f = %f" %(calc.fck, calc.fy, calc.ρvmin)
        if calc.ρvuse >= calc.ρvmin :
            wsout['C86'].value ="pv,use = %4.1f / (%4.1f x %4.1f x sin%3.1f) = %1.6f ≥ %1.6f ∴ O.K" %(calc.Avs, calc.AvSpace, calc.B, calc.αv, calc.ρvuse, calc.ρvmin)
        else :
            wsout['C86'].value ="pv,use = %4.1f / (%4.1f x %4.1f x sin%3.1f) = %1.6f < %1.6f ∴ N.G" %(calc.Avs, calc.AvSpace, calc.B, calc.αv, calc.ρvuse, calc.ρvmin)
        wsout['B88'].value ="12) 전단철근 간격검토"
        if calc.s1max >= calc.AvSpace :
            wsout['C89'].value ="- 종방향 Smax = 0.75 x %4.1f x (1 + cot%3.1f) = %4.1f ≥ %4.1f   ∴ O.K" %(calc.D, calc.αv, calc.s1max, calc.AvSpace)
        else :
            wsout['C89'].value ="- 종방향 Smax = 0.75 x %4.1f x (1 + cot%3.1f) = %4.1f < %4.1f   ∴ N.G" %(calc.D, calc.αv, calc.s1max, calc.AvSpace)    
        if calc.s2max >= calc.s2 :
            wsout['C90'].value ="- 횡방향 Smax = min(0.75 x %4.1f, 600.0) = %4.1f ≥ %4.1f   ∴ O.K" %(calc.D, calc.s2max, calc.s2)
        else :
            wsout['C90'].value ="- 횡방향 Smax = min(0.75 x %4.1f, 600.0) = %4.1f < %4.1f   ∴ N.G" %(calc.D, calc.s2max, calc.s2) 
        wsout['B92'].value ="13) 작용전단력에 의해 종방향 철근에 발생하는 추가인장력 검토"
        wsout['C93'].value ="ΔTr = (Mr - Mu) / z = (%10.1f - %10.1f) / %4.1f = %9.1f N" %(calc.Mr, calc.Mun, calc.D, calc.dTr)
        wsout['C94'].value ="ΔT = 0.5 x Vu x (cotθ - cotα) = 0.5 x %9.1f x (cot%3.1f - cot%3.1f)" %(calc.Vun, calc.θ, calc.αv)
        if calc.dT >= calc.dTr :
            wsout['C95'].value ="     = %9.1f N ≥ %9.1f N  → ∴ 추가 주철근 필요" %(calc.dT, calc.dTr)
        else :
            wsout['C95'].value ="     = %9.1f N < %9.1f N  → ∴ 추가 주철근 필요 없음" %(calc.dT, calc.dTr)


def serviceexcelout(calc, wsout) :
    wsout['B100'].value ="14) 균열검토"
    wsout['B101'].value =" - 한계상태 검증을 위한 설계등급 : E (표면 한계균열폭 0.3 mm)"
    wsout['B102'].value =" ① 비균열 가정시 인장 연단 응력(사용하중조합-Ⅰ)"
    wsout['C103'].value ="ft = M / Zb = M / (b x h² / 6))"
    wsout['C104'].value ="   = %10.1f / (%4.1f x %4.1f² / 6))" %(calc.Msn1, calc.B, calc.H)
    if calc.fct1 < calc.fctk :
        wsout['C105'].value ="   = %10.1f mm² < %4.1f mm²  → ∴ 비균열단면으로 균열검토 필요없음" %(calc.fct1, calc.fctk)
    else :
        wsout['C105'].value ="   = %10.1f mm² ≥ %4.1f mm²  → ∴ 균열단면으로 균열검토 필요" %(calc.fct1, calc.fctk)
        wsout['B106'].value =" ② 균열응력 검토"
        wsout['C107'].value ="n = Es / Ec = %d" %(calc.nr)
        wsout['C108'].value ="c = -n x As / b + n x As / b x √(1 + 2 x b x d / (n x As))"
        wsout['C109'].value ="  = -%d x %4.1f / %4.1f + %d x %6.1f / %4.1f x √(1 + 2 x %4.1f x %4.1f / (%d x %4.1f))" %(calc.nr, calc.Asuse, calc.B, calc.nr, calc.Asuse, calc.B, calc.B, calc.D, calc.nr, calc.Asuse)
        wsout['C110'].value ="  = %2.3f mm" %(calc.c)
        wsout['C111'].value ="- 콘크리트 연단에 발생하는 응력" 
        wsout['D112'].value ="fc1 = 2 x M1 / ((1 - k / 3) x b x c x d)"
        wsout['D113'].value ="    = 2 x %10.0f / ((1 - %4.1f / 3) x %4.1f x %2.3f x %4.1f)" %(calc.Msn1, calc.k, calc.B, calc.c, calc.D)
        if calc.fc1 < calc.fca1 :
            wsout['D114'].value ="    = %4.3f N/mm² < 0.6 x fck = %4.1f N/mm²   ∴ O.k" %(calc.fc1, calc.fca1)
        else :
            wsout['D114'].value ="    = %4.3f N/mm² ≥ 0.6 x fck = %4.1f N/mm²   ∴ N.G" %(calc.fc1, calc.fca1)
        wsout['D115'].value ="fc5 = 2 x M5 / ((1 - k / 3) x b x c x d)"
        wsout['D116'].value ="    = 2 x %10.0f / ((1 - %4.1f / 3) x %4.1f x %2.3f x %4.1f)" %(calc.Msn5, calc.k, calc.B, calc.c, calc.D)
        if calc.fc1 < calc.fca1 :
            wsout['D117'].value ="    = %4.3f N/mm² < 0.45 x fck = %4.1f N/mm²   ∴ O.k" %(calc.fc5, calc.fca5)
        else :
            wsout['D117'].value ="    = %4.3f N/mm² ≥ 0.45 x fck = %4.1f N/mm²   ∴ N.G" %(calc.fc5, calc.fca5)
        wsout['C118'].value ="- 철근에 발생하는 응력" 
        wsout['D119'].value =" fs1 = M1 / (As x (1 - k / 3) x d)"
        wsout['D120'].value ="     = %10.1f / (%6.1f x (1 - %4.1f / 3) x %4.1f)" %(calc.Msn1, calc.Asuse, calc.k, calc.D)
        if calc.fs1 < calc.fsa :
            wsout['D121'].value ="     = %4.3f N/mm² < %4.1f N/mm²   ∴ O.k" %(calc.fs1, calc.fsa)
        else :
            wsout['D121'].value ="     = %4.3f N/mm² ≥ %4.1f N/mm²   ∴ N.G" %(calc.fs1, calc.fsa)
        wsout['B123'].value =" ③ 균열제어를 위한 최소철근량"
        wsout['C124'].value ="As,min = kc x k x Act x (fct / fs) [도.설(2015) 5.8.3.2]" 
        wsout['C125'].value ="       = %2.3f x %2.3f x %6.1f x (%2.3f / %4.1f) = %6.1fmm²" %(calc.kc, calc.k, calc.Act, calc.fct, calc.fsa,calc.Asdmin) 
        wsout['C126'].value =" - Act = min[2.5 x (h - d), (h - x)/3, h/2] x B"
        wsout['C127'].value ="       = min[2.5 x (%4.1f - %4.1f), (%4.1f - %4.1f)/3, %4.1f/2] x %4.1f" %(calc.H, calc.D, calc.H, calc.x, calc.H, calc.B)
        wsout['C128'].value ="       = %6.1f mm²" %(calc.Act)
        wsout['C129'].value =" -  x = - n x As / b + n x As / b x √(1 + 2 x b x d / (n x As))"
        wsout['C130'].value ="      = - %d x %6.1f / %4.1f + %d x %6.1f / %4.1f x √(1 + 2 x %4.1f x %4.1f / (%d x %6.1f))" %(calc.nr, calc.Asuse, calc.B, calc.nr, calc.Asuse, calc.B, calc.B, calc.D, calc.nr, calc.Asuse)
        wsout['C131'].value ="      = %3.2f mm" %(calc.x)
        if calc.H <= 300:    
            wsout['C132'].value =" -  k = %2.3f ( H ≤ 300 )" %(calc.k)
        elif calc.H < 800:
            wsout['C132'].value =" -  k = -0.0007 x %3.1f + 1.21 = %2.3f ( H < 800 )" %(calc.H, calc.k)
        else :
            wsout['C132'].value =" -  k = %2.3f ( H ≥ 800 )" %(calc.k)
        wsout['C133'].value =" -  kc = 0.4 x [1 - fn / (k1 x (h / h') x fct)] ≤ 1"
        wsout['C134'].value ="       = 0.4 x [1 - %3.3f / (%1.3f x (%4.1f / %4.1f) x %2.3f)]" %(calc.fnn, calc.k1, calc.H, calc.h1, calc.fct)
        if calc.Asuse >= calc.Asdmin :
            wsout['C135'].value ="As,use = %6.1f mm² ≥ As,min = %6.1f mm²    ∴ O.K" %(calc.Asuse, calc.Asdmin)
        else :
            wsout['C135'].value ="As,use = %6.1f mm² < As,min = %6.1f mm²    ∴ N.G" %(calc.Asuse, calc.Asdmin)
        wsout['B137'].value =" ④ 간접균열검토"
        wsout['C138'].value =" - fs5 = Ms5 / (As x (1 - k / 3) x d)"
        wsout['C139'].value ="       = %10.1f / (%6.1f x (1 - %4.1f / 3) x %4.1f) = %4.3f N/mm²" %(calc.Msn5, calc.Asuse, calc.k, calc.D, calc.fs5)
        wsout['C140'].value =" - 사용철근 직경 : D = %dmm, fs = %4.1fN/mm²" %(calc.AsDia1, calc.fsad)
        wsout['C141'].value =" - 사용철근 간격 : s = %dmm, fs = %4.1fN/mm²" %(calc.Asspace1, calc.fsas)
        if calc.fsaf >= calc.fs5 :
            wsout['C142'].value =" - 철근 한계응력 : MAX(%4.1f, %4.1f) ≥ fs = %4.1f N/mm²  ∴ O.K" %(calc.fsad, calc.fsas, calc.fs5)
        else :
            wsout['C142'].value =" - 철근 한계응력 : MAX(%4.1f, %4.1f) < fs = %4.1f N/mm²  ∴ N.G" %(calc.fsad, calc.fsas, calc.fs5)


#--------------------------------------------------
#    직사각형 단철근보 단면검토 (도로교설계기준 2012)
#--------------------------------------------------

def oldfunc():
    #wb = load_workbook("D:/Python/Excel/Calc_As_input_3.0.xlsx",read_only=False)
    fname = filedialog.askopenfile(title="Open input",filetypes=[("Excel files", "*.xlsx")]).name
    wb = load_workbook(fname,read_only=False)

    wsin = wb['Sheet1']
    data_rc_mat = []  #재료특성 입력 data_rc_mat
    data_mat_fac_uls = [] #재료계수(극한한계상태) 입력 data_matfac_uls 
    data_mat_fac_als = [] #재료계수(극단상황한계상태) 입력 data_matfac_als
    data_sec = [] #단면제원 입력 data_sec
    data_usedas_band = [] #사용휨철근량 입력 data_usedas_band
    data_usedas_shear = [] #전단철근량 및 각도 입력 data_usedas_shear
    data_secforce_uls = [] #산정하중(극한한계상태) 입력 data_secforce_uls
    data_secforce_als = [] #산정하중(극단상황한계상태) 입력 data_secforce_als
    datatp = wsin['C4:D4']                   #재료특성 입력
    for datatp1 in datatp :                  #튜플형을 리스트형으로 변환 
        for datatp2 in datatp1 :
            data_rc_mat.append(datatp2.value)
    datatp3 = wsin['C7:D7']                   #재료계수(극한한계상태) 입력
    print("datap3:",list(datatp3)) 
    for datatp4 in datatp3 :                  #튜플형을 리스트형으로 변환 
        for datatp5 in datatp4 :
            data_mat_fac_uls.append(datatp5.value)
    datatp6 = wsin['C8:D8']                   #재료계수(극단상황한계상태) 입력
    for datatp7 in datatp6 :                  #튜플형을 리스트형으로 변환 
        for datatp8 in datatp7 :
            data_mat_fac_als.append(datatp8.value)
    datatp9 = wsin['C11:D11']                 #단면제원 입력
    for datatp10 in datatp9 :                 #튜플형을 리스트형으로 변환 
        for datatp11 in datatp10 :
            data_sec.append(datatp11.value)
    datatp12 = wsin['C15:K15']                 #사용휨철근량 입력  
    for datatp13 in datatp12 :                 #튜플형을 리스트형으로 변환 
        for datatp14 in datatp13 :
            data_usedas_band.append(datatp14.value)
    datatp15 = wsin['C18:H18']                 #전단철근량 및 각도 입력
    for datatp16 in datatp15 :                 #튜플형을 리스트형으로 변환 
        for datatp17 in datatp16 :
            data_usedas_shear.append(datatp17.value)
    datatp18 = wsin['C21:H21']                 #산정하중(극한한계상태) 입력
    for datatp19 in datatp18 :                 #튜플형을 리스트형으로 변환 
        for datatp20 in datatp19 :
            data_secforce_uls.append(datatp20.value)
    datatp21 = wsin['C22:H22']                 #산정하중(극단상황한계상태) 입력
    for datatp22 in datatp21 :                 #튜플형을 리스트형으로 변환 
        for datatp23 in datatp22 :
            data_secforce_als.append(datatp23.value)
    calc = Sec_back(data_rc_mat, data_mat_fac_uls, data_sec, data_usedas_band, data_usedas_shear, data_secforce_uls)
    calc1 = Sec_back(data_rc_mat, data_matfac_als, data_sec, data_usedas_band, data_usedas_shear, data_secforce_als)
    wb.close()

    #------------------------------------
    #          휨모멘트 검토 결과 출력
    #------------------------------------
    calc.calmoment()
    calc1.calmoment()
    wb = Workbook()
    wsout = wb.active
    wsout.title = "극한한계상태"
    wsout1 = wb.create_sheet("극단상황한계상태")

    def wsinit(wsout) :                             #워크시트 형식 초기화
        alpalist = list(ascii_uppercase)
        for i in range(1,100) :                     # 헹 크기 지정
            wsout.row_dimensions[i].height = 15
        for i in alpalist :                         # 열 크기 지정
            wsout.column_dimensions[i].width = 3.0
        font_format = Font(size=9, name = '굴림체')
        for rows in wsout["A1":"Z200"] :            # 기본 폰트를 size=9, 굴림체로 변경
            for cell in rows :
                cell.font = font_format

    def mexcelsheet(calc, wsout) :
        wsout['B2'].value = '1) 단면제원 및 설계가정'
        wsout['C3'].value ="fck = %d Mpa, fy = %d Mpa, Øc = %1.2f, Øs = %1.2f, Es = %d Mpa" %(calc.fck, calc.fy, calc.Øc, calc.Øs, calc.Es)
        for r in range(4,6) :                        # 4~5행 글짜 위치 중앙으로 정렬
            for c in range(3,24) :                      
                wsout.range(r,c).alignment = Alignment(horizontal='center', vertical='center')  
        for r in range(4,6) :                        # 4~5행 테두리 그리기
            for c in range(3,24) :                      
                wsout.range(r,c).border = Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        for r in range(4,6) :                        # 셀 병합 C4~L4, C5~L5
            for c in [3,6,9,12] :                       
                c1 = c + 2
                #wsout.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c1)
                wsout.range((r,c), (r, c1)).merge()
        for r in range(4,6) :                         # 셀 병합 O4~T4, O5~T5  
            for c in [15,20] :                          
                c1 = c + 4
                #wsout.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c1) 
                wsout.range((r, c), (r,c1)).merge()    
        for c in range(3,24) :                         # 셀 채우기 C4~T5
            wsout.range(4,c).fill = PatternFill(fill_type='solid', fgColor='0000FFFF')
        wsout['C4'].value = 'B(mm)'
        wsout['F4'].value = 'H(mm)'
        wsout['I4'].value = 'd(mm)'
        wsout['L4'].value = '피복(mm)'
        wsout['O4'].value = 'Mu(N.mm)'
        wsout['T4'].value = 'Vu(N)'
        wsout['C5'].value = calc.B
        wsout['F5'].value = calc.H
        wsout['I5'].value = "%4.3f" %(calc.D)
        wsout['L5'].value = "%4.3f" %(calc.Dc)
        wsout['O5'].value = calc.Mun
        wsout['O5'].number_format = '#,##0#'
        wsout['T5'].value = calc.Vun
        wsout['T5'].number_format = '#,##0#'
        wsout['B7'].value = '2) 콘크리트 재료상수'
        for r in range(8,19) :
            #wsout.merge_cells(start_row=r, start_column=17, end_row=r, end_column=18)
            wsout.range((r,17),(r,18)).merge()
        wsout['C8'].value = 'n      : 상승 곡선부의 형상을 나타내는 지수'
        wsout['P8'].value = '='
        wsout['Q8'].value = calc.nε
        wsout['C9'].value = 'εco,r : 최대응력에 처음 도달했을때의 변형률'
        wsout['P9'].value = '='
        wsout['Q9'].value = calc.εco
        wsout['C10'].value = 'εcu,r : 극한변형률'
        wsout['P10'].value = '='
        wsout['Q10'].value = calc.εcu
        wsout['C11'].value = 'αcc   : 유효계수'
        wsout['P11'].value = '='
        wsout['Q11'].value = calc.αcc
        wsout['C12'].value = 'fcd   : 콘크리트 설계압축강도'
        wsout['P12'].value = '='
        wsout['Q12'].value = calc.fcd
        wsout['S12'].value = 'MPa'
        wsout['C13'].value = 'fcm   : 평균압축강도(fck+Δf)'
        wsout['P13'].value = '='
        wsout['Q13'].value = calc.fcm
        wsout['S13'].value = 'MPa'
        wsout['C14'].value = 'Ec    : 콘크리트 탄성계수'
        wsout['P14'].value = '='
        wsout['Q14'].value = calc.Ec
        wsout['S14'].value = 'MPa'
        wsout['C15'].value = 'α     : 압축합력의 평균 응력계수'
        wsout['P15'].value = '='
        wsout['Q15'].value = calc.α
        wsout['C16'].value = 'β     : 압축합력의 작용점 위치계수'
        wsout['P16'].value = '='
        wsout['Q16'].value = calc.β
        wsout['C17'].value = 'η     : 등가 사각형 응력 블록의 크기계수'
        wsout['P17'].value = '='
        wsout['Q17'].value = calc.η
        wsout['C18'].value = 'β1    : 등가 사각형 응력 블록의 깊이계수(2β)'
        wsout['P18'].value = '='
        wsout['Q18'].value = calc.β1

        wsout['B20'].value = '3) 철근 재료상수'
        for r in range(21,23) :
            #wsout.merge_cells(start_row=r, start_column=17, end_row=r, end_column=18)
            wsout.range((r, 17), (r,18)).merge()
        wsout['C21'].value = 'fyd    : 설계인장강도 ( Φs fy )'
        wsout['P21'].value = '='
        wsout['Q21'].value = calc.fyd
        wsout['S21'].value = 'MPa'
        wsout['C22'].value = 'εyd    : 설계 항복 변형률 ( fyd / Es )'
        wsout['P22'].value = '='
        wsout['Q22'].value = calc.εyd

        wsout['B24'].value = '4) 필요철근량 산정'
        wsout['C25'].value = 'Mu = As x fyd  x (d - a / 2)              ----------------   ①'
        wsout['C26'].value = ' a = As x fyd  / ( η x fcd x b)          ----------------   ②'
        wsout['C27'].value = ' 식②를 식①에 대입하여 이차방정식으로 As를 구한다'
        wsout['F28'].value = ' fyd²'
        wsout['C29'].value = ' ──────────── As² - fyd x d x As + Mu = 0 ,   Asreq = %6.1fmm²' %(calc.Asreq)
        wsout['D30'].value = ' 2 x η x fcd x b'

        wsout['B32'].value ="5) 사용철근량 : Asuse =  %6.1f mm², 철근도심 : dc =  %5.1f mm [ 사용율 = %3.3f ]" %(calc.Asuse, calc.Dc, calc.Asuse/calc.Asreq)
        wsout['F33'].value = "1단 : %c %d - %d EA (=  %6.1f mm², dc1 =  %4.1f mm)" %(calc.rebarid, calc.AsDia1, calc.AsNum1, calc.Asuse1*calc.AsNum1, calc.Dc1)
        wsout['F34'].value = "2단 : %c %d - %d EA (=  %6.1f mm², dc2 =  %4.1f mm)" %(calc.rebarid, calc.AsDia2, calc.AsNum2, calc.Asuse2*calc.AsNum2, calc.Dc2)
        wsout['F35'].value = "3단 : %c %d - %d EA (=  %6.1f mm², dc3 =  %4.1f mm)" %(calc.rebarid, calc.AsDia3, calc.AsNum3, calc.Asuse3*calc.AsNum3, calc.Dc3)

        wsout['B37'].value ="6) 철근량 검토"
        wsout['C38'].value ="As,min = (0.25 √fck / fy) x b x d = %6.1f mm²" %(calc.Asmin1)
        wsout['D39'].value ="   = (1.4 / fy) x b x d = %6.1f mm²" %(calc.Asmin2) 
        wsout['D40'].value ="   = As_req x 4 / 3 = %6.1f mm²" %(calc.Asmin3) 
        if calc.Asuse >= calc.Asmin :
            wsout['C41'].value ="As,use = %6.1f mm² ≥ As,min = %6.1f mm²  ∴ O.K" %(calc.Asuse, calc.Asmin)
        else :
            wsout['C41'].value ="As,use = %6.1f mm² < As,min = %6.1f mm²  ∴ N.G" %(calc.Asuse, calc.Asmin)
        if calc.Asuse <= calc.Asmax :
            wsout['C42'].value ="As,max = 0.04 x b x d = %6.1f mm² ≥  As,use = %6.1f mm²  ∴ O.K" %(calc.Asmax, calc.Asuse)
        else :
            wsout['C42'].value ="As,max = 0.04 x b x d = %6.1f mm² <  As,use = %6.1f mm²  ∴ N.G" %(calc.Asmax, calc.Asuse)

        wsout['B44'].value ="7) 중립축 깊이 검토"
        wsout['C45'].value ="Cmax = (δ x εcu / 0.0033 - 0.6) x d" 
        wsout['D46'].value =" = (%2.1f x %2.5f / 0.0033 - 0.6) x %4.1f = %4.1f mm" %(calc.δ, calc.εcu, calc.D, calc.c_max)
        wsout['C47'].value ="C = Φs x As x fy / (α x Φc x 0.85 x fck x b)" 
        wsout['C48'].value ="  = %2.2f x %6.1f x %d / (%2.2f x %2.2f x 0.85 x %d x %5.1f)" %(calc.Øs, calc.Asuse, calc.fy, calc.α, calc.Øc, calc.fck, calc.B)
        if calc.cc <= calc.c_max :
            wsout['C49'].value ="  = %6.1f mm < Cmax = %6.1f mm  ∴ O.K" %(calc.cc, calc.c_max)
        else :
            wsout['C49'].value ="  = %6.1f mm ≥ Cmax = %6.1f mm  ∴ N.G" %(calc.cc, calc.c_max)

        wsout['B51'].value ="8) 인장철근 변형률"
        wsout['C52'].value ="εs = (d - C) / C x εcu" 
        wsout['C53'].value ="    = (%4.1f - %4.1f) / %4.1f x %2.5f = %2.5f " %(calc.D, calc.cc, calc.cc, calc.εcu, calc.εs)
        wsout['C54'].value ="εyd = Φs x fy / Es" 
        if calc.εyd <= calc.εs :
            wsout['D55'].value =" = %2.2f x %d / %d = %2.5f  ≤ εs  ∴ 항복가정 O.K" %(calc.Øs, calc.fy, calc.Es, calc.εyd)
        else :
            wsout['D55'].value =" = %2.2f x %d / %d = %2.5f  > εs  ∴ 항복가정 N.G" %(calc.Øs, calc.fy, calc.Es, calc.εyd)

        wsout['B57'].value ="9) 설계 휨강도 산정"
        wsout['C58'].value ="Mr = As x Φs x fy x (d - β x c)" 
        wsout['C59'].value ="   = %6.1f x %2.2f x %d x (%4.1f - %2.2f x %4.1f)" %(calc.Asuse, calc.Øs, calc.fy, calc.D, calc.β, calc.cc)
        if calc.Mr > calc.Mun :
            wsout['C60'].value ="   = %10.1f N.mm  ≥ Mu = %10.1f N.mm  ∴ O.K  [S.F = %3.3f]" %(calc.Mr, calc.Mun, calc.Msf)
        else :
            wsout['C60'].value ="   = %10.1f N.mm  < Mu = %10.1f N.mm  ∴ N.G  [S.F = %3.3f]" %(calc.Mr, calc.Mun, calc.Msf)
    wsinit(wsout)
    wsinit(wsout1)
    mexcelsheet(calc, wsout)
    mexcelsheet(calc1, wsout1)
    #------------------------------
    #          전단력 검토
    #------------------------------
    calc.calshear()
    calc1.calshear()
    def sexcelsheet(calc, wsout) :
        wsout['B62'].value ="10) 전단검토"
        wsout['C63'].value ="Vcd = [0.85 x Φc x k x (p x fck)⅓ + 0.15 x fn] x b x d" 
        wsout['C64'].value ="    = [0.85 x %1.2f x %1.3f x (%1.6f x %3.1f)⅓ + 0.15 x %4.3f] x %4.1f x %4.1f" %(calc.Øc, calc.k, calc.ρs, calc.fck, calc.fn, calc.B, calc.D)
        wsout['C65'].value ="    = %10.0f N" %(calc.Vc) 
        wsout['D66'].value ="k = 1 + √(200 / d) = 1 + √(200 / %4.1f) = %1.3f (≤ 2.000)" %(calc.D, calc.k)
        wsout['D67'].value ="p = As / (b x d) = %6.1f / (%4.1f x %4.1f) = %1.6f ≤ 0.0200" %(calc.Asuse, calc.B, calc.D, calc.ρ) 
        wsout['D68'].value ="fn = Nu / Ac = %4.3f MPa (≤ 0.2 Φc fck = %4.3f MPa, 압축의경우+)" %(calc.fnn, calc.fnmax)
        wsout['C69'].value ="Vcd,min = (0.4 x Φc x fctk + 0.15 x fn] x b x d" 
        wsout['C70'].value ="        = (0.4 x %1.2f x %3.3f + 0.15 x %4.3f] x %4.1f x %4.1f" %(calc.Øc, calc.fctk, calc.fn, calc.B, calc.D)
        wsout['C71'].value ="        = %9.0f N" %(calc.Vcdmin)

        if calc.Vcd >= calc.Vun :
            wsout['C72'].value ="Vcd = %9.0f N ≥ Vu = %9.0f N  ∴ 전단보강 불필요" %(calc.Vcd, calc.Vun)
            wsout['B74'].value ="11) 최소 전단철근량 검토"
            wsout['C75'].value ="Av,use = %5.1f mm² ( %c %d - %2.3fea, C.T.C = %4.1f mm )" %(calc.Avs, calc.rebarid, calc.AvDia, calc.AvLeg, calc.AvSpace)
            wsout['C76'].value ="pv,min = (0.08 x √fck) / fvy = (0.08 x √%3.1f) / %4.1f = %1.6f" %(calc.fck, calc.fy, calc.ρvmin)
            if calc.ρvuse >= calc.ρvmin :
                wsout['C77'].value ="pv,use = %4.1f / (%4.1f x %4.1f x sin%3.1f) = %1.6f ≥ %1.6f ∴ O.K" %(calc.Avs, calc.AvSpace, calc.B, calc.αv, calc.ρvuse, calc.ρvmin)
            else :
                wsout['C77'].value ="pv,use = %4.1f / (%4.1f x %4.1f x sin%3.1f) = %1.6f < %1.6f ∴ N.G" %(calc.Avs, calc.AvSpace, calc.B, calc.αv, calc.ρvuse, calc.ρvmin)
            wsout['B79'].value ="12) 전단철근 간격검토"
            if calc.s1max >= calc.AvSpace :
                wsout['C80'].value ="- 종방향 Smax = 0.75 x %4.1f x (1 + cot%3.1f) = %4.1f ≥ %4.1f   ∴ O.K" %(calc.D, calc.αv, calc.s1max, calc.AvSpace)
            else :
                wsout['C80'].value ="- 종방향 Smax = 0.75 x %4.1f x (1 + cot%3.1f) = %4.1f < %4.1f   ∴ N.G" %(calc.D, calc.αv, calc.s1max, calc.AvSpace)    
            if calc.s2max >= calc.s2 :
                wsout['C81'].value ="- 횡방향 Smax = min(0.75 x %4.1f, 600.0) = %4.1f ≥ %4.1f   ∴ O.K" %(calc.D, calc.s2max, calc.s2)
            else :
                wsout['C81'].value ="- 종방향 Smax = min(0.75 x %4.1f, 600.0) = %4.1f < %4.1f   ∴ N.G" %(calc.D, calc.s2max, calc.s2)
        else :
            wsout['C72'].value ="Vcd = %9.0f N < Vu = %9.0f N  ∴ 전단보강 필요" %(calc.Vcd, calc.Vun)
            wsout['C74'].value ="cotΘ = %1.3f ( Θ = %3.1f˚)" %(calc.cotθ, calc.θ)
            if calc.cotθ < 1 :
                wsout['L74'].value ="1.0 > cotΘ → Θ 불만족  ∴ N.G"
            elif calc.cotθ > 2.5 :    
                wsout['L74'].value ="2.5 ≤ cotΘ → Θ 불만족  ∴ N.G"
            else :
                wsout['L74'].value ="1.0 ≤ cotΘ ≤ 2.5 → Θ 만족  ∴ O.K"
            wsout['C76'].value ="Av,use = %5.1f mm² ( %c %d - %2.3fea, C.T.C = %4.1f mm )" %(calc.Avs, calc.rebarid, calc.AvDia, calc.AvLeg, calc.AvSpace)
            wsout['C77'].value ="Vsd = (Φs x fvy x Av x z / s) x cotθ"
            wsout['C78'].value ="    = (%1.2f x %4.1f x %5.1f x %4.1f / %4.1f) x %1.3f" %(calc.Øs, calc.fy, calc.Avs, calc.z, calc.AvSpace, calc.cotθ)
            if calc.Vd >= calc.Vun :    
                wsout['C79'].value ="    = %d N ≥ Vu = %d N  ∴ O.K" %(calc.Vd, calc.Vun)
            else :
                wsout['C79'].value ="    = %d N < Vu = %d N  ∴ N.G" %(calc.Vd, calc.Vun)    
            wsout['C80'].value ="Vsd,max = (v x Φc x fck x b x z) / (cotθ + tanθ)"
            wsout['C81'].value ="        = (%1.3f x %1.2f x %3.1f x %4.1f x %4.1f) / ( %1.3f + %1.3f )" %(calc.ν, calc.Øc, calc.fck, calc.B, calc.z, calc.cotθ, calc.tanθ)
            if calc.Vd <= calc.Vdmax :
                wsout['C82'].value ="        = %d N ≤ Vsd = %d N  ∴ O.K" %(calc.Vdmax, calc.Vd)
            else :
                wsout['C82'].value ="        = %d N > Vsd = %d N  ∴ N.G" %(calc.Vdmax, calc.Vd)
            wsout['B84'].value ="11) 최소 전단철근량 검토"
            wsout['C85'].value ="pv,min = (0.08 x √fck) / fvy = (0.08 x √%3.1f) / %4.1f = %f" %(calc.fck, calc.fy, calc.ρvmin)
            if calc.ρvuse >= calc.ρvmin :
                wsout['C86'].value ="pv,use = %4.1f / (%4.1f x %4.1f x sin%3.1f) = %1.6f ≥ %1.6f ∴ O.K" %(calc.Avs, calc.AvSpace, calc.B, calc.αv, calc.ρvuse, calc.ρvmin)
            else :
                wsout['C86'].value ="pv,use = %4.1f / (%4.1f x %4.1f x sin%3.1f) = %1.6f < %1.6f ∴ N.G" %(calc.Avs, calc.AvSpace, calc.B, calc.αv, calc.ρvuse, calc.ρvmin)
            wsout['B88'].value ="12) 전단철근 간격검토"
            if calc.s1max >= calc.AvSpace :
                wsout['C89'].value ="- 종방향 Smax = 0.75 x %4.1f x (1 + cot%3.1f) = %4.1f ≥ %4.1f   ∴ O.K" %(calc.D, calc.αv, calc.s1max, calc.AvSpace)
            else :
                wsout['C89'].value ="- 종방향 Smax = 0.75 x %4.1f x (1 + cot%3.1f) = %4.1f < %4.1f   ∴ N.G" %(calc.D, calc.αv, calc.s1max, calc.AvSpace)    
            if calc.s2max >= calc.s2 :
                wsout['C90'].value ="- 횡방향 Smax = min(0.75 x %4.1f, 600.0) = %4.1f ≥ %4.1f   ∴ O.K" %(calc.D, calc.s2max, calc.s2)
            else :
                wsout['C90'].value ="- 횡방향 Smax = min(0.75 x %4.1f, 600.0) = %4.1f < %4.1f   ∴ N.G" %(calc.D, calc.s2max, calc.s2) 
            wsout['B92'].value ="13) 작용전단력에 의해 종방향 철근에 발생하는 추가인장력 검토"
            wsout['C93'].value ="ΔTr = (Mr - Mu) / z = (%10.1f - %10.1f) / %4.1f = %9.1f N" %(calc.Mr, calc.Mun, calc.D, calc.dTr)
            wsout['C94'].value ="ΔT = 0.5 x Vu x (cotθ - cotα) = 0.5 x %9.1f x (cot%3.1f - cot%3.1f)" %(calc.Vun, calc.θ, calc.αv)
            if calc.dT >= calc.dTr :
                wsout['C95'].value ="     = %9.1f N ≥ %9.1f N  → ∴ 추가 주철근 필요" %(calc.dT, calc.dTr)
            else :
                wsout['C95'].value ="     = %9.1f N < %9.1f N  → ∴ 추가 주철근 필요 없음" %(calc.dT, calc.dTr)

    sexcelsheet(calc, wsout)
    sexcelsheet(calc1, wsout1)
    #--------------------------------------
    #          사용성 검토(균열검토)
    #--------------------------------------
    calc.calservice()
    def vexcelsheet(calc, wsout) :
        wsout['B100'].value ="14) 균열검토"
        wsout['B101'].value =" - 한계상태 검증을 위한 설계등급 : E (표면 한계균열폭 0.3 mm)"
        wsout['B102'].value =" ① 비균열 가정시 인장 연단 응력(사용하중조합-Ⅰ)"
        wsout['C103'].value ="ft = M / Zb = M / (b x h² / 6))"
        wsout['C104'].value ="   = %10.1f / (%4.1f x %4.1f² / 6))" %(calc.Msn1, calc.B, calc.H)
        if calc.fct1 < calc.fctk :
            wsout['C105'].value ="   = %10.1f mm² < %4.1f mm²  → ∴ 비균열단면으로 균열검토 필요없음" %(calc.fct1, calc.fctk)
        else :
            wsout['C105'].value ="   = %10.1f mm² ≥ %4.1f mm²  → ∴ 균열단면으로 균열검토 필요" %(calc.fct1, calc.fctk)
            wsout['B106'].value =" ② 균열응력 검토"
            wsout['C107'].value ="n = Es / Ec = %d" %(calc.nr)
            wsout['C108'].value ="c = -n x As / b + n x As / b x √(1 + 2 x b x d / (n x As))"
            wsout['C109'].value ="  = -%d x %4.1f / %4.1f + %d x %6.1f / %4.1f x √(1 + 2 x %4.1f x %4.1f / (%d x %4.1f))" %(calc.nr, calc.Asuse, calc.B, calc.nr, calc.Asuse, calc.B, calc.B, calc.D, calc.nr, calc.Asuse)
            wsout['C110'].value ="  = %2.3f mm" %(calc.c)
            wsout['C111'].value ="- 콘크리트 연단에 발생하는 응력" 
            wsout['D112'].value ="fc1 = 2 x M1 / ((1 - k / 3) x b x c x d)"
            wsout['D113'].value ="    = 2 x %10.0f / ((1 - %4.1f / 3) x %4.1f x %2.3f x %4.1f)" %(calc.Msn1, calc.k, calc.B, calc.c, calc.D)
            if calc.fc1 < calc.fca1 :
                wsout['D114'].value ="    = %4.3f N/mm² < 0.6 x fck = %4.1f N/mm²   ∴ O.k" %(calc.fc1, calc.fca1)
            else :
                wsout['D114'].value ="    = %4.3f N/mm² ≥ 0.6 x fck = %4.1f N/mm²   ∴ N.G" %(calc.fc1, calc.fca1)
            wsout['D115'].value ="fc5 = 2 x M5 / ((1 - k / 3) x b x c x d)"
            wsout['D116'].value ="    = 2 x %10.0f / ((1 - %4.1f / 3) x %4.1f x %2.3f x %4.1f)" %(calc.Msn5, calc.k, calc.B, calc.c, calc.D)
            if calc.fc1 < calc.fca1 :
                wsout['D117'].value ="    = %4.3f N/mm² < 0.45 x fck = %4.1f N/mm²   ∴ O.k" %(calc.fc5, calc.fca5)
            else :
                wsout['D117'].value ="    = %4.3f N/mm² ≥ 0.45 x fck = %4.1f N/mm²   ∴ N.G" %(calc.fc5, calc.fca5)
            wsout['C118'].value ="- 철근에 발생하는 응력" 
            wsout['D119'].value =" fs1 = M1 / (As x (1 - k / 3) x d)"
            wsout['D120'].value ="     = %10.1f / (%6.1f x (1 - %4.1f / 3) x %4.1f)" %(calc.Msn1, calc.Asuse, calc.k, calc.D)
            if calc.fs1 < calc.fsa :
                wsout['D121'].value ="     = %4.3f N/mm² < %4.1f N/mm²   ∴ O.k" %(calc.fs1, calc.fsa)
            else :
                wsout['D121'].value ="     = %4.3f N/mm² ≥ %4.1f N/mm²   ∴ N.G" %(calc.fs1, calc.fsa)
            wsout['B123'].value =" ③ 균열제어를 위한 최소철근량"
            wsout['C124'].value ="As,min = kc x k x Act x (fct / fs) [도.설(2015) 5.8.3.2]" 
            wsout['C125'].value ="       = %2.3f x %2.3f x %6.1f x (%2.3f / %4.1f) = %6.1fmm²" %(calc.kc, calc.k, calc.Act, calc.fct, calc.fsa,calc.Asdmin) 
            wsout['C126'].value =" - Act = min[2.5 x (h - d), (h - x)/3, h/2] x B"
            wsout['C127'].value ="       = min[2.5 x (%4.1f - %4.1f), (%4.1f - %4.1f)/3, %4.1f/2] x %4.1f" %(calc.H, calc.D, calc.H, calc.x, calc.H, calc.B)
            wsout['C128'].value ="       = %6.1f mm²" %(calc.Act)
            wsout['C129'].value =" -  x = - n x As / b + n x As / b x √(1 + 2 x b x d / (n x As))"
            wsout['C130'].value ="      = - %d x %6.1f / %4.1f + %d x %6.1f / %4.1f x √(1 + 2 x %4.1f x %4.1f / (%d x %6.1f))" %(calc.nr, calc.Asuse, calc.B, calc.nr, calc.Asuse, calc.B, calc.B, calc.D, calc.nr, calc.Asuse)
            wsout['C131'].value ="      = %3.2f mm" %(calc.x)
            if calc.H <= 300:    
                wsout['C132'].value =" -  k = %2.3f ( H ≤ 300 )" %(calc.k)
            elif calc.H < 800:
                wsout['C132'].value =" -  k = -0.0007 x %3.1f + 1.21 = %2.3f ( H < 800 )" %(calc.H, calc.k)
            else :
                wsout['C132'].value =" -  k = %2.3f ( H ≥ 800 )" %(calc.k)
            wsout['C133'].value =" -  kc = 0.4 x [1 - fn / (k1 x (h / h') x fct)] ≤ 1"
            wsout['C134'].value ="       = 0.4 x [1 - %3.3f / (%1.3f x (%4.1f / %4.1f) x %2.3f)]" %(calc.fnn, calc.k1, calc.H, calc.h1, calc.fct)
            if calc.Asuse >= calc.Asdmin :
                wsout['C135'].value ="As,use = %6.1f mm² ≥ As,min = %6.1f mm²    ∴ O.K" %(calc.Asuse, calc.Asdmin)
            else :
                wsout['C135'].value ="As,use = %6.1f mm² < As,min = %6.1f mm²    ∴ N.G" %(calc.Asuse, calc.Asdmin)
            wsout['B137'].value =" ④ 간접균열검토"
            wsout['C138'].value =" - fs5 = Ms5 / (As x (1 - k / 3) x d)"
            wsout['C139'].value ="       = %10.1f / (%6.1f x (1 - %4.1f / 3) x %4.1f) = %4.3f N/mm²" %(calc.Msn5, calc.Asuse, calc.k, calc.D, calc.fs5)
            wsout['C140'].value =" - 사용철근 직경 : D = %dmm, fs = %4.1fN/mm²" %(calc.AsDia1, calc.fsad)
            wsout['C141'].value =" - 사용철근 간격 : s = %dmm, fs = %4.1fN/mm²" %(calc.Asspace1, calc.fsas)
            if calc.fsaf >= calc.fs5 :
                wsout['C142'].value =" - 철근 한계응력 : MAX(%4.1f, %4.1f) ≥ fs = %4.1f N/mm²  ∴ O.K" %(calc.fsad, calc.fsas, calc.fs5)
            else :
                wsout['C142'].value =" - 철근 한계응력 : MAX(%4.1f, %4.1f) < fs = %4.1f N/mm²  ∴ N.G" %(calc.fsad, calc.fsas, calc.fs5)

    vexcelsheet(calc, wsout)   
    wfname = filedialog.asksaveasfile(title="Save result",filetypes=[("Excel files", "*.xlsx")]).name
    wb.save(wfname)
    #wb.save('Calc_As_Output_3.0.xlsx')
    wb.close()

#oldfunc()
#readxls()

# ============================================
# 텍스트 기반 계산과정 출력 함수들 (GUI 뷰어용)
# ============================================

def get_moment_text(calc):
    """휨모멘트 계산과정을 텍스트로 반환"""
    lines = []
    lines.append("=" * 60)
    lines.append("1) 단면제원 및 설계가정")
    lines.append("=" * 60)
    lines.append(f"fck = {calc.fck} MPa, fy = {calc.fy} MPa, Øc = {calc.Øc:.2f}, Øs = {calc.Øs:.2f}, Es = {calc.Es} MPa")
    lines.append("")
    lines.append(f"  B = {calc.B:.1f} mm")
    lines.append(f"  H = {calc.H:.1f} mm")
    lines.append(f"  d = {calc.D:.1f} mm")
    lines.append(f"  피복 = {calc.Dc:.1f} mm")
    lines.append(f"  Mu = {calc.Mun:,.0f} N.mm")
    lines.append(f"  Vu = {calc.Vun:,.0f} N")
    lines.append("")

    lines.append("=" * 60)
    lines.append("2) 콘크리트 재료상수")
    lines.append("=" * 60)
    lines.append(f"  n (상승곡선부 형상지수)      = {calc.nε:.4f}")
    lines.append(f"  εco,r (최대응력 변형률)      = {calc.εco:.5f}")
    lines.append(f"  εcu,r (극한변형률)           = {calc.εcu:.5f}")
    lines.append(f"  αcc (유효계수)               = {calc.αcc}")
    lines.append(f"  fcd (설계압축강도)           = {calc.fcd:.2f} MPa")
    lines.append(f"  fcm (평균압축강도)           = {calc.fcm:.2f} MPa")
    lines.append(f"  Ec (탄성계수)                = {calc.Ec:.2f} MPa")
    lines.append(f"  α (압축합력 응력계수)        = {calc.α:.4f}")
    lines.append(f"  β (압축합력 위치계수)        = {calc.β:.4f}")
    lines.append(f"  η (등가응력블록 크기계수)    = {calc.η:.4f}")
    lines.append(f"  β1 (등가응력블록 깊이계수)   = {calc.β1:.4f}")
    lines.append("")

    lines.append("=" * 60)
    lines.append("3) 철근 재료상수")
    lines.append("=" * 60)
    lines.append(f"  fyd (설계인장강도)           = {calc.fyd:.2f} MPa")
    lines.append(f"  εyd (설계항복변형률)         = {calc.εyd:.6f}")
    lines.append("")

    lines.append("=" * 60)
    lines.append("4) 필요철근량 산정")
    lines.append("=" * 60)
    lines.append("  Mu = As × fyd × (d - a/2)  -------- ①")
    lines.append("   a = As × fyd / (η × fcd × b)  ---- ②")
    lines.append("  식②를 식①에 대입하여 이차방정식으로 As를 구한다")
    lines.append("")
    lines.append(f"  Asreq = {calc.Asreq:.1f} mm²")
    lines.append("")

    lines.append("=" * 60)
    lines.append("5) 사용철근량")
    lines.append("=" * 60)
    lines.append(f"  Asuse = {calc.Asuse:.1f} mm², 철근도심: dc = {calc.Dc:.1f} mm")
    lines.append(f"  사용율 = {calc.Asuse/calc.Asreq:.3f}")
    lines.append(f"  1단: {calc.rebarid}{calc.AsDia1} - {calc.AsNum1} EA (= {calc.Asuse1*calc.AsNum1:.1f} mm², dc1 = {calc.Dc1:.1f} mm)")
    lines.append(f"  2단: {calc.rebarid}{calc.AsDia2} - {calc.AsNum2} EA (= {calc.Asuse2*calc.AsNum2:.1f} mm², dc2 = {calc.Dc2:.1f} mm)")
    lines.append(f"  3단: {calc.rebarid}{calc.AsDia3} - {calc.AsNum3} EA (= {calc.Asuse3*calc.AsNum3:.1f} mm², dc3 = {calc.Dc3:.1f} mm)")
    lines.append("")

    lines.append("=" * 60)
    lines.append("6) 철근량 검토")
    lines.append("=" * 60)
    lines.append(f"  As,min = (0.25√fck / fy) × b × d = {calc.Asmin1:.1f} mm²")
    lines.append(f"         = (1.4 / fy) × b × d = {calc.Asmin2:.1f} mm²")
    lines.append(f"         = As_req × 4/3 = {calc.Asmin3:.1f} mm²")
    if calc.Asuse >= calc.Asmin:
        lines.append(f"  As,use = {calc.Asuse:.1f} mm² ≥ As,min = {calc.Asmin:.1f} mm²  ∴ O.K")
    else:
        lines.append(f"  As,use = {calc.Asuse:.1f} mm² < As,min = {calc.Asmin:.1f} mm²  ∴ N.G")
    if calc.Asuse <= calc.Asmax:
        lines.append(f"  As,max = 0.04 × b × d = {calc.Asmax:.1f} mm² ≥ As,use  ∴ O.K")
    else:
        lines.append(f"  As,max = 0.04 × b × d = {calc.Asmax:.1f} mm² < As,use  ∴ N.G")
    lines.append("")

    lines.append("=" * 60)
    lines.append("7) 중립축 깊이 검토")
    lines.append("=" * 60)
    lines.append(f"  Cmax = (δ × εcu / 0.0033 - 0.6) × d")
    lines.append(f"       = ({calc.δ:.1f} × {calc.εcu:.5f} / 0.0033 - 0.6) × {calc.D:.1f} = {calc.c_max:.1f} mm")
    lines.append(f"  C = Φs × As × fy / (α × Φc × 0.85 × fck × b)")
    lines.append(f"    = {calc.Øs:.2f} × {calc.Asuse:.1f} × {calc.fy} / ({calc.α:.2f} × {calc.Øc:.2f} × 0.85 × {calc.fck} × {calc.B:.1f})")
    if calc.cc <= calc.c_max:
        lines.append(f"    = {calc.cc:.1f} mm < Cmax = {calc.c_max:.1f} mm  ∴ O.K")
    else:
        lines.append(f"    = {calc.cc:.1f} mm ≥ Cmax = {calc.c_max:.1f} mm  ∴ N.G")
    lines.append("")

    lines.append("=" * 60)
    lines.append("8) 인장철근 변형률")
    lines.append("=" * 60)
    lines.append(f"  εs = (d - C) / C × εcu")
    lines.append(f"     = ({calc.D:.1f} - {calc.cc:.1f}) / {calc.cc:.1f} × {calc.εcu:.5f} = {calc.εs:.5f}")
    lines.append(f"  εyd = Φs × fy / Es")
    if calc.εyd <= calc.εs:
        lines.append(f"      = {calc.Øs:.2f} × {calc.fy} / {calc.Es} = {calc.εyd:.5f} ≤ εs  ∴ 항복가정 O.K")
    else:
        lines.append(f"      = {calc.Øs:.2f} × {calc.fy} / {calc.Es} = {calc.εyd:.5f} > εs  ∴ 항복가정 N.G")
    lines.append("")

    lines.append("=" * 60)
    lines.append("9) 설계 휨강도 산정")
    lines.append("=" * 60)
    lines.append(f"  Mr = As × Φs × fy × (d - β × c)")
    lines.append(f"     = {calc.Asuse:.1f} × {calc.Øs:.2f} × {calc.fy} × ({calc.D:.1f} - {calc.β:.2f} × {calc.cc:.1f})")
    if calc.Mr > calc.Mun:
        lines.append(f"     = {calc.Mr:,.1f} N.mm ≥ Mu = {calc.Mun:,.1f} N.mm  ∴ O.K")
    else:
        lines.append(f"     = {calc.Mr:,.1f} N.mm < Mu = {calc.Mun:,.1f} N.mm  ∴ N.G")
    lines.append(f"  [S.F = {calc.Msf:.3f}]")
    lines.append("")

    return "\n".join(lines)


def get_shear_text(calc):
    """전단력 계산과정을 텍스트로 반환"""
    lines = []
    lines.append("=" * 60)
    lines.append("10) 전단검토")
    lines.append("=" * 60)
    lines.append(f"  Vcd = [0.85 × Φc × k × (ρ × fck)^(1/3) + 0.15 × fn] × b × d")
    lines.append(f"      = [0.85 × {calc.Øc:.2f} × {calc.k:.3f} × ({calc.ρs:.6f} × {calc.fck:.1f})^(1/3) + 0.15 × {calc.fn:.3f}] × {calc.B:.1f} × {calc.D:.1f}")
    lines.append(f"      = {calc.Vc:,.0f} N")
    lines.append("")
    lines.append(f"  k = 1 + √(200/d) = 1 + √(200/{calc.D:.1f}) = {calc.k:.3f} (≤ 2.000)")
    lines.append(f"  ρ = As / (b × d) = {calc.Asuse:.1f} / ({calc.B:.1f} × {calc.D:.1f}) = {calc.ρ:.6f} ≤ 0.0200")
    lines.append(f"  fn = Nu / Ac = {calc.fnn:.3f} MPa (≤ 0.2Φc×fck = {calc.fnmax:.3f} MPa)")
    lines.append("")
    lines.append(f"  Vcd,min = (0.4 × Φc × fctk + 0.15 × fn) × b × d")
    lines.append(f"         = (0.4 × {calc.Øc:.2f} × {calc.fctk:.3f} + 0.15 × {calc.fn:.3f}) × {calc.B:.1f} × {calc.D:.1f}")
    lines.append(f"         = {calc.Vcdmin:,.0f} N")
    lines.append("")

    if calc.Vcd >= calc.Vun:
        lines.append(f"  Vcd = {calc.Vcd:,.0f} N ≥ Vu = {calc.Vun:,.0f} N  ∴ 전단보강 불필요")
        lines.append("")
        lines.append("=" * 60)
        lines.append("11) 최소 전단철근량 검토")
        lines.append("=" * 60)
        lines.append(f"  Av,use = {calc.Avs:.1f} mm² ({calc.rebarid}{calc.AvDia} - {calc.AvLeg}ea, C.T.C = {calc.AvSpace:.1f} mm)")
        lines.append(f"  ρv,min = (0.08 × √fck) / fvy = (0.08 × √{calc.fck:.1f}) / {calc.fy:.1f} = {calc.ρvmin:.6f}")
        if calc.ρvuse >= calc.ρvmin:
            lines.append(f"  ρv,use = {calc.Avs:.1f} / ({calc.AvSpace:.1f} × {calc.B:.1f} × sin{calc.αv:.1f}) = {calc.ρvuse:.6f} ≥ {calc.ρvmin:.6f}  ∴ O.K")
        else:
            lines.append(f"  ρv,use = {calc.Avs:.1f} / ({calc.AvSpace:.1f} × {calc.B:.1f} × sin{calc.αv:.1f}) = {calc.ρvuse:.6f} < {calc.ρvmin:.6f}  ∴ N.G")
        lines.append("")
        lines.append("=" * 60)
        lines.append("12) 전단철근 간격검토")
        lines.append("=" * 60)
        if calc.s1max >= calc.AvSpace:
            lines.append(f"  종방향 Smax = 0.75 × {calc.D:.1f} × (1 + cot{calc.αv:.1f}) = {calc.s1max:.1f} ≥ {calc.AvSpace:.1f}  ∴ O.K")
        else:
            lines.append(f"  종방향 Smax = 0.75 × {calc.D:.1f} × (1 + cot{calc.αv:.1f}) = {calc.s1max:.1f} < {calc.AvSpace:.1f}  ∴ N.G")
        if calc.s2max >= calc.s2:
            lines.append(f"  횡방향 Smax = min(0.75 × {calc.D:.1f}, 600.0) = {calc.s2max:.1f} ≥ {calc.s2:.1f}  ∴ O.K")
        else:
            lines.append(f"  횡방향 Smax = min(0.75 × {calc.D:.1f}, 600.0) = {calc.s2max:.1f} < {calc.s2:.1f}  ∴ N.G")
    else:
        lines.append(f"  Vcd = {calc.Vcd:,.0f} N < Vu = {calc.Vun:,.0f} N  ∴ 전단보강 필요")
        lines.append("")
        lines.append(f"  cotθ = {calc.cotθ:.3f} (θ = {calc.θ:.1f}°)")
        if calc.cotθ < 1:
            lines.append("  1.0 > cotθ → θ 불만족  ∴ N.G")
        elif calc.cotθ > 2.5:
            lines.append("  2.5 < cotθ → θ 불만족  ∴ N.G")
        else:
            lines.append("  1.0 ≤ cotθ ≤ 2.5 → θ 만족  ∴ O.K")
        lines.append("")
        lines.append(f"  Av,use = {calc.Avs:.1f} mm² ({calc.rebarid}{calc.AvDia} - {calc.AvLeg}ea, C.T.C = {calc.AvSpace:.1f} mm)")
        lines.append(f"  Vsd = (Φs × fvy × Av × z / s) × cotθ")
        lines.append(f"      = ({calc.Øs:.2f} × {calc.fy:.1f} × {calc.Avs:.1f} × {calc.z:.1f} / {calc.AvSpace:.1f}) × {calc.cotθ:.3f}")
        if calc.Vd >= calc.Vun:
            lines.append(f"      = {calc.Vd:,.0f} N ≥ Vu = {calc.Vun:,.0f} N  ∴ O.K")
        else:
            lines.append(f"      = {calc.Vd:,.0f} N < Vu = {calc.Vun:,.0f} N  ∴ N.G")
        lines.append("")
        lines.append(f"  Vsd,max = (ν × Φc × fck × b × z) / (cotθ + tanθ)")
        lines.append(f"         = ({calc.ν:.3f} × {calc.Øc:.2f} × {calc.fck:.1f} × {calc.B:.1f} × {calc.z:.1f}) / ({calc.cotθ:.3f} + {calc.tanθ:.3f})")
        if calc.Vd <= calc.Vdmax:
            lines.append(f"         = {calc.Vdmax:,.0f} N ≥ Vsd = {calc.Vd:,.0f} N  ∴ O.K")
        else:
            lines.append(f"         = {calc.Vdmax:,.0f} N < Vsd = {calc.Vd:,.0f} N  ∴ N.G")
        lines.append("")
        lines.append("=" * 60)
        lines.append("11) 최소 전단철근량 검토")
        lines.append("=" * 60)
        lines.append(f"  ρv,min = (0.08 × √fck) / fvy = (0.08 × √{calc.fck:.1f}) / {calc.fy:.1f} = {calc.ρvmin:.6f}")
        if calc.ρvuse >= calc.ρvmin:
            lines.append(f"  ρv,use = {calc.ρvuse:.6f} ≥ {calc.ρvmin:.6f}  ∴ O.K")
        else:
            lines.append(f"  ρv,use = {calc.ρvuse:.6f} < {calc.ρvmin:.6f}  ∴ N.G")
        lines.append("")
        lines.append("=" * 60)
        lines.append("12) 전단철근 간격검토")
        lines.append("=" * 60)
        if calc.s1max >= calc.AvSpace:
            lines.append(f"  종방향 Smax = {calc.s1max:.1f} ≥ {calc.AvSpace:.1f}  ∴ O.K")
        else:
            lines.append(f"  종방향 Smax = {calc.s1max:.1f} < {calc.AvSpace:.1f}  ∴ N.G")
        if calc.s2max >= calc.s2:
            lines.append(f"  횡방향 Smax = {calc.s2max:.1f} ≥ {calc.s2:.1f}  ∴ O.K")
        else:
            lines.append(f"  횡방향 Smax = {calc.s2max:.1f} < {calc.s2:.1f}  ∴ N.G")
        lines.append("")
        lines.append("=" * 60)
        lines.append("13) 추가인장력 검토")
        lines.append("=" * 60)
        lines.append(f"  ΔTr = (Mr - Mu) / z = ({calc.Mr:,.1f} - {calc.Mun:,.1f}) / {calc.D:.1f} = {calc.dTr:,.1f} N")
        lines.append(f"  ΔT = 0.5 × Vu × (cotθ - cotα) = 0.5 × {calc.Vun:,.1f} × (cot{calc.θ:.1f} - cot{calc.αv:.1f})")
        if calc.dT >= calc.dTr:
            lines.append(f"     = {calc.dT:,.1f} N ≥ {calc.dTr:,.1f} N  → ∴ 추가 주철근 필요")
        else:
            lines.append(f"     = {calc.dT:,.1f} N < {calc.dTr:,.1f} N  → ∴ 추가 주철근 필요 없음")

    lines.append("")
    return "\n".join(lines)


def get_service_text(calc):
    """사용성 검토(균열검토) 과정을 텍스트로 반환"""
    lines = []
    lines.append("=" * 60)
    lines.append("14) 균열검토")
    lines.append("=" * 60)
    lines.append(" - 한계상태 검증을 위한 설계등급 : E (표면 한계균열폭 0.3 mm)")
    lines.append("")
    lines.append("① 비균열 가정시 인장 연단 응력(사용하중조합-I)")
    lines.append(f"  ft = M / Zb = M / (b × h² / 6)")
    lines.append(f"     = {calc.Msn1:,.1f} / ({calc.B:.1f} × {calc.H:.1f}² / 6)")
    if calc.fct1 < calc.fctk:
        lines.append(f"     = {calc.fct1:.1f} MPa < {calc.fctk:.1f} MPa  → ∴ 비균열단면으로 균열검토 필요없음")
    else:
        lines.append(f"     = {calc.fct1:.1f} MPa ≥ {calc.fctk:.1f} MPa  → ∴ 균열단면으로 균열검토 필요")
        lines.append("")
        lines.append("② 균열응력 검토")
        lines.append(f"  n = Es / Ec = {calc.nr}")
        lines.append(f"  c = -n×As/b + n×As/b × √(1 + 2×b×d/(n×As))")
        lines.append(f"    = {calc.c:.3f} mm")
        lines.append("")
        lines.append("  - 콘크리트 연단에 발생하는 응력")
        lines.append(f"    fc1 = 2 × M1 / ((1 - k/3) × b × c × d)")
        lines.append(f"        = 2 × {calc.Msn1:,.0f} / ((1 - {calc.k:.1f}/3) × {calc.B:.1f} × {calc.c:.3f} × {calc.D:.1f})")
        if calc.fc1 < calc.fca1:
            lines.append(f"        = {calc.fc1:.3f} N/mm² < 0.6×fck = {calc.fca1:.1f} N/mm²  ∴ O.K")
        else:
            lines.append(f"        = {calc.fc1:.3f} N/mm² ≥ 0.6×fck = {calc.fca1:.1f} N/mm²  ∴ N.G")
        lines.append("")
        lines.append(f"    fc5 = 2 × M5 / ((1 - k/3) × b × c × d)")
        lines.append(f"        = 2 × {calc.Msn5:,.0f} / ((1 - {calc.k:.1f}/3) × {calc.B:.1f} × {calc.c:.3f} × {calc.D:.1f})")
        if calc.fc5 < calc.fca5:
            lines.append(f"        = {calc.fc5:.3f} N/mm² < 0.45×fck = {calc.fca5:.1f} N/mm²  ∴ O.K")
        else:
            lines.append(f"        = {calc.fc5:.3f} N/mm² ≥ 0.45×fck = {calc.fca5:.1f} N/mm²  ∴ N.G")
        lines.append("")
        lines.append("  - 철근에 발생하는 응력")
        lines.append(f"    fs1 = M1 / (As × (1 - k/3) × d)")
        lines.append(f"        = {calc.Msn1:,.1f} / ({calc.Asuse:.1f} × (1 - {calc.k:.1f}/3) × {calc.D:.1f})")
        if calc.fs1 < calc.fsa:
            lines.append(f"        = {calc.fs1:.3f} N/mm² < {calc.fsa:.1f} N/mm²  ∴ O.K")
        else:
            lines.append(f"        = {calc.fs1:.3f} N/mm² ≥ {calc.fsa:.1f} N/mm²  ∴ N.G")
        lines.append("")
        lines.append("③ 균열제어를 위한 최소철근량")
        lines.append(f"  As,min = kc × k × Act × (fct / fs) [도.설(2015) 5.8.3.2]")
        lines.append(f"         = {calc.kc:.3f} × {calc.k:.3f} × {calc.Act:.1f} × ({calc.fct:.3f} / {calc.fsa:.1f}) = {calc.Asdmin:.1f} mm²")
        lines.append(f"  - Act = {calc.Act:.1f} mm²")
        lines.append(f"  - x = {calc.x:.2f} mm")
        lines.append(f"  - k = {calc.k:.3f}")
        lines.append(f"  - kc = {calc.kc:.3f}")
        if calc.Asuse >= calc.Asdmin:
            lines.append(f"  As,use = {calc.Asuse:.1f} mm² ≥ As,min = {calc.Asdmin:.1f} mm²  ∴ O.K")
        else:
            lines.append(f"  As,use = {calc.Asuse:.1f} mm² < As,min = {calc.Asdmin:.1f} mm²  ∴ N.G")
        lines.append("")
        lines.append("④ 간접균열검토")
        lines.append(f"  - fs5 = Ms5 / (As × (1 - k/3) × d)")
        lines.append(f"        = {calc.Msn5:,.1f} / ({calc.Asuse:.1f} × (1 - {calc.k:.1f}/3) × {calc.D:.1f}) = {calc.fs5:.3f} N/mm²")
        lines.append(f"  - 사용철근 직경: D = {calc.AsDia1}mm, fs = {calc.fsad:.1f} N/mm²")
        lines.append(f"  - 사용철근 간격: s = {calc.Asspace1:.0f}mm, fs = {calc.fsas:.1f} N/mm²")
        if calc.fsaf >= calc.fs5:
            lines.append(f"  - 철근 한계응력: MAX({calc.fsad:.1f}, {calc.fsas:.1f}) ≥ fs = {calc.fs5:.1f} N/mm²  ∴ O.K")
        else:
            lines.append(f"  - 철근 한계응력: MAX({calc.fsad:.1f}, {calc.fsas:.1f}) < fs = {calc.fs5:.1f} N/mm²  ∴ N.G")

    lines.append("")
    return "\n".join(lines)


def get_full_calc_text(calc):
    """전체 계산과정을 텍스트로 반환"""
    text = get_moment_text(calc)
    text += get_shear_text(calc)
    text += get_service_text(calc)
    return text
