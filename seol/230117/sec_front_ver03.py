#import xlwings as xw
import openpyxl as op
from openpyxl import load_workbook, Workbook
from string import ascii_uppercase
from openpyxl.styles.fonts import Font
from openpyxl.styles import Border, Side, Alignment, PatternFill
from openpyxl.styles.colors import Color
from sec_back_ver03 import *
from tkinter import filedialog

#--------------------------------------------------
#    직사각형 단철근보 단면검토 (도로교설계기준 2012)
#--------------------------------------------------

#wb = load_workbook("D:/Python/Excel/Calc_As_input_3.0.xlsx",read_only=False)
fname = filedialog.askopenfile(title="Open input",filetypes=[("Excel files", "*.xlsx")]).name
wb = load_workbook(fname,read_only=False)

wsin = wb['Sheet1']
datalist = []  #재료특성 입력 data_rc_mat
datalist1 = [] #재료계수(극한한계상태) 입력 data_matfac_uls 
datalist2 = [] #재료계수(극단상황한계상태) 입력 data_matfac_als
datalist3 = [] #단면제원 입력 data_sec
datalist4 = [] #사용휨철근량 입력 data_usedas_band
datalist5 = [] #전단철근량 및 각도 입력 data_usedas_shear
datalist6 = [] #산정하중(극한한계상태) 입력 data_secforce_uls
datalist7 = [] #산정하중(극단상황한계상태) 입력 data_secforce_als
datatp = wsin['C4:D4']                   #재료특성 입력
for datatp1 in datatp :                  #튜플형을 리스트형으로 변환 
    for datatp2 in datatp1 :
        datalist.append(datatp2.value)
datatp3 = wsin['C7:D7']                   #재료계수(극한한계상태) 입력
print("datap3:",list(datatp3)) 
for datatp4 in datatp3 :                  #튜플형을 리스트형으로 변환 
    for datatp5 in datatp4 :
        datalist1.append(datatp5.value)
datatp6 = wsin['C8:D8']                   #재료계수(극단상황한계상태) 입력
for datatp7 in datatp6 :                  #튜플형을 리스트형으로 변환 
    for datatp8 in datatp7 :
        datalist2.append(datatp8.value)
datatp9 = wsin['C11:D11']                 #단면제원 입력
for datatp10 in datatp9 :                 #튜플형을 리스트형으로 변환 
    for datatp11 in datatp10 :
        datalist3.append(datatp11.value)
datatp12 = wsin['C15:K15']                 #사용휨철근량 입력  
for datatp13 in datatp12 :                 #튜플형을 리스트형으로 변환 
    for datatp14 in datatp13 :
        datalist4.append(datatp14.value)
datatp15 = wsin['C18:H18']                 #전단철근량 및 각도 입력
for datatp16 in datatp15 :                 #튜플형을 리스트형으로 변환 
    for datatp17 in datatp16 :
        datalist5.append(datatp17.value)
datatp18 = wsin['C21:H21']                 #산정하중(극한한계상태) 입력
for datatp19 in datatp18 :                 #튜플형을 리스트형으로 변환 
    for datatp20 in datatp19 :
        datalist6.append(datatp20.value)
datatp21 = wsin['C22:H22']                 #산정하중(극단상황한계상태) 입력
for datatp22 in datatp21 :                 #튜플형을 리스트형으로 변환 
    for datatp23 in datatp22 :
        datalist7.append(datatp23.value)
calc = Sec_back(datalist, datalist1, datalist3, datalist4, datalist5, datalist6)
calc1 = Sec_back(datalist, datalist2, datalist3, datalist4, datalist5, datalist7)
wb.close()

#------------------------------------
#          휨모멘트 검토 결과 출력
#------------------------------------
calc.calmoment()
calc1.calmoment()
wb = Workbook()
wsout = wb.active
wsout.title = "극한한계상태"
wsout1 = wb.create_sheet("극단상황한계상태")

def wsinit(wsout) :                             #워크시트 형식 초기화
    alpalist = list(ascii_uppercase)
    for i in range(1,100) :                     # 헹 크기 지정
        wsout.row_dimensions[i].height = 15
    for i in alpalist :                         # 열 크기 지정
        wsout.column_dimensions[i].width = 3.0
    font_format = Font(size=9, name = '굴림체')
    for rows in wsout["A1":"Z200"] :            # 기본 폰트를 size=9, 굴림체로 변경
        for cell in rows :
            cell.font = font_format

def mexcelsheet(calc, wsout) :
    wsout['B2'].value = '1) 단면제원 및 설계가정'
    wsout['C3'].value ="fck = %d Mpa, fy = %d Mpa, Øc = %1.2f, Øs = %1.2f, Es = %d Mpa" %(calc.fck, calc.fy, calc.Øc, calc.Øs, calc.Es)
    for r in range(4,6) :                        # 4~5행 글짜 위치 중앙으로 정렬
        for c in range(3,24) :                      
            wsout.cell(r,c).alignment = Alignment(horizontal='center', vertical='center')  
    for r in range(4,6) :                        # 4~5행 테두리 그리기
        for c in range(3,24) :                      
            wsout.cell(r,c).border = Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    for r in range(4,6) :                        # 셀 병합 C4~L4, C5~L5
        for c in [3,6,9,12] :                       
            c1 = c + 2
            wsout.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c1)
    for r in range(4,6) :                         # 셀 병합 O4~T4, O5~T5  
        for c in [15,20] :                          
            c1 = c + 4
            wsout.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c1)    
    for c in range(3,24) :                         # 셀 채우기 C4~T5
        wsout.cell(4,c).fill = PatternFill(fill_type='solid', fgColor='0000FFFF')
    wsout['C4'].value = 'B(mm)'
    wsout['F4'].value = 'H(mm)'
    wsout['I4'].value = 'd(mm)'
    wsout['L4'].value = '피복(mm)'
    wsout['O4'].value = 'Mu(N.mm)'
    wsout['T4'].value = 'Vu(N)'
    wsout['C5'].value = calc.B
    wsout['F5'].value = calc.H
    wsout['I5'].value = "%4.3f" %(calc.D)
    wsout['L5'].value = "%4.3f" %(calc.Dc)
    wsout['O5'].value = calc.Mun
    wsout['O5'].number_format = '#,##0#'
    wsout['T5'].value = calc.Vun
    wsout['T5'].number_format = '#,##0#'
    wsout['B7'].value = '2) 콘크리트 재료상수'
    for r in range(8,19) :
        wsout.merge_cells(start_row=r, start_column=17, end_row=r, end_column=18)
    wsout['C8'].value = 'n      : 상승 곡선부의 형상을 나타내는 지수'
    wsout['P8'].value = '='
    wsout['Q8'].value = calc.nε
    wsout['C9'].value = 'εco,r : 최대응력에 처음 도달했을때의 변형률'
    wsout['P9'].value = '='
    wsout['Q9'].value = calc.εco
    wsout['C10'].value = 'εcu,r : 극한변형률'
    wsout['P10'].value = '='
    wsout['Q10'].value = calc.εcu
    wsout['C11'].value = 'αcc   : 유효계수'
    wsout['P11'].value = '='
    wsout['Q11'].value = calc.αcc
    wsout['C12'].value = 'fcd   : 콘크리트 설계압축강도'
    wsout['P12'].value = '='
    wsout['Q12'].value = calc.fcd
    wsout['S12'].value = 'MPa'
    wsout['C13'].value = 'fcm   : 평균압축강도(fck+Δf)'
    wsout['P13'].value = '='
    wsout['Q13'].value = calc.fcm
    wsout['S13'].value = 'MPa'
    wsout['C14'].value = 'Ec    : 콘크리트 탄성계수'
    wsout['P14'].value = '='
    wsout['Q14'].value = calc.Ec
    wsout['S14'].value = 'MPa'
    wsout['C15'].value = 'α     : 압축합력의 평균 응력계수'
    wsout['P15'].value = '='
    wsout['Q15'].value = calc.α
    wsout['C16'].value = 'β     : 압축합력의 작용점 위치계수'
    wsout['P16'].value = '='
    wsout['Q16'].value = calc.β
    wsout['C17'].value = 'η     : 등가 사각형 응력 블록의 크기계수'
    wsout['P17'].value = '='
    wsout['Q17'].value = calc.η
    wsout['C18'].value = 'β1    : 등가 사각형 응력 블록의 깊이계수(2β)'
    wsout['P18'].value = '='
    wsout['Q18'].value = calc.β1

    wsout['B20'].value = '3) 철근 재료상수'
    for r in range(21,23) :
        wsout.merge_cells(start_row=r, start_column=17, end_row=r, end_column=18)
    wsout['C21'].value = 'fyd    : 설계인장강도 ( Φs fy )'
    wsout['P21'].value = '='
    wsout['Q21'].value = calc.fyd
    wsout['S21'].value = 'MPa'
    wsout['C22'].value = 'εyd    : 설계 항복 변형률 ( fyd / Es )'
    wsout['P22'].value = '='
    wsout['Q22'].value = calc.εyd

    wsout['B24'].value = '4) 필요철근량 산정'
    wsout['C25'].value = 'Mu = As x fyd  x (d - a / 2)              ----------------   ①'
    wsout['C26'].value = ' a = As x fyd  / ( η x fcd x b)          ----------------   ②'
    wsout['C27'].value = ' 식②를 식①에 대입하여 이차방정식으로 As를 구한다'
    wsout['F28'].value = ' fyd²'
    wsout['C29'].value = ' ──────────── As² - fyd x d x As + Mu = 0 ,   Asreq = %6.1fmm²' %(calc.Asreq)
    wsout['D30'].value = ' 2 x η x fcd x b'

    wsout['B32'].value ="5) 사용철근량 : Asuse =  %6.1f mm², 철근도심 : dc =  %5.1f mm [ 사용율 = %3.3f ]" %(calc.Asuse, calc.Dc, calc.Asuse/calc.Asreq)
    wsout['F33'].value = "1단 : %c %d - %d EA (=  %6.1f mm², dc1 =  %4.1f mm)" %(calc.rebarid, calc.AsDia1, calc.AsNum1, calc.Asuse1*calc.AsNum1, calc.Dc1)
    wsout['F34'].value = "2단 : %c %d - %d EA (=  %6.1f mm², dc2 =  %4.1f mm)" %(calc.rebarid, calc.AsDia2, calc.AsNum2, calc.Asuse2*calc.AsNum2, calc.Dc2)
    wsout['F35'].value = "3단 : %c %d - %d EA (=  %6.1f mm², dc3 =  %4.1f mm)" %(calc.rebarid, calc.AsDia3, calc.AsNum3, calc.Asuse3*calc.AsNum3, calc.Dc3)

    wsout['B37'].value ="6) 철근량 검토"
    wsout['C38'].value ="As,min = (0.25 √fck / fy) x b x d = %6.1f mm²" %(calc.Asmin1)
    wsout['D39'].value ="   = (1.4 / fy) x b x d = %6.1f mm²" %(calc.Asmin2) 
    wsout['D40'].value ="   = As_req x 4 / 3 = %6.1f mm²" %(calc.Asmin3) 
    if calc.Asuse >= calc.Asmin :
        wsout['C41'].value ="As,use = %6.1f mm² ≥ As,min = %6.1f mm²  ∴ O.K" %(calc.Asuse, calc.Asmin)
    else :
        wsout['C41'].value ="As,use = %6.1f mm² < As,min = %6.1f mm²  ∴ N.G" %(calc.Asuse, calc.Asmin)
    if calc.Asuse <= calc.Asmax :
        wsout['C42'].value ="As,max = 0.04 x b x d = %6.1f mm² ≥  As,use = %6.1f mm²  ∴ O.K" %(calc.Asmax, calc.Asuse)
    else :
        wsout['C42'].value ="As,max = 0.04 x b x d = %6.1f mm² <  As,use = %6.1f mm²  ∴ N.G" %(calc.Asmax, calc.Asuse)

    wsout['B44'].value ="7) 중립축 깊이 검토"
    wsout['C45'].value ="Cmax = (δ x εcu / 0.0033 - 0.6) x d" 
    wsout['D46'].value =" = (%2.1f x %2.5f / 0.0033 - 0.6) x %4.1f = %4.1f mm" %(calc.δ, calc.εcu, calc.D, calc.c_max)
    wsout['C47'].value ="C = Φs x As x fy / (α x Φc x 0.85 x fck x b)" 
    wsout['C48'].value ="  = %2.2f x %6.1f x %d / (%2.2f x %2.2f x 0.85 x %d x %5.1f)" %(calc.Øs, calc.Asuse, calc.fy, calc.α, calc.Øc, calc.fck, calc.B)
    if calc.cc <= calc.c_max :
        wsout['C49'].value ="  = %6.1f mm < Cmax = %6.1f mm  ∴ O.K" %(calc.cc, calc.c_max)
    else :
        wsout['C49'].value ="  = %6.1f mm ≥ Cmax = %6.1f mm  ∴ N.G" %(calc.cc, calc.c_max)

    wsout['B51'].value ="8) 인장철근 변형률"
    wsout['C52'].value ="εs = (d - C) / C x εcu" 
    wsout['C53'].value ="    = (%4.1f - %4.1f) / %4.1f x %2.5f = %2.5f " %(calc.D, calc.cc, calc.cc, calc.εcu, calc.εs)
    wsout['C54'].value ="εyd = Φs x fy / Es" 
    if calc.εyd <= calc.εs :
        wsout['D55'].value =" = %2.2f x %d / %d = %2.5f  ≤ εs  ∴ 항복가정 O.K" %(calc.Øs, calc.fy, calc.Es, calc.εyd)
    else :
        wsout['D55'].value =" = %2.2f x %d / %d = %2.5f  > εs  ∴ 항복가정 N.G" %(calc.Øs, calc.fy, calc.Es, calc.εyd)

    wsout['B57'].value ="9) 설계 휨강도 산정"
    wsout['C58'].value ="Mr = As x Φs x fy x (d - β x c)" 
    wsout['C59'].value ="   = %6.1f x %2.2f x %d x (%4.1f - %2.2f x %4.1f)" %(calc.Asuse, calc.Øs, calc.fy, calc.D, calc.β, calc.cc)
    if calc.Mr > calc.Mun :
        wsout['C60'].value ="   = %10.1f N.mm  ≥ Mu = %10.1f N.mm  ∴ O.K  [S.F = %3.3f]" %(calc.Mr, calc.Mun, calc.Msf)
    else :
        wsout['C60'].value ="   = %10.1f N.mm  < Mu = %10.1f N.mm  ∴ N.G  [S.F = %3.3f]" %(calc.Mr, calc.Mun, calc.Msf)
wsinit(wsout)
wsinit(wsout1)
mexcelsheet(calc, wsout)
mexcelsheet(calc1, wsout1)
#------------------------------
#          전단력 검토
#------------------------------
calc.calshear()
calc1.calshear()
def sexcelsheet(calc, wsout) :
    wsout['B62'].value ="10) 전단검토"
    wsout['C63'].value ="Vcd = [0.85 x Φc x k x (p x fck)⅓ + 0.15 x fn] x b x d" 
    wsout['C64'].value ="    = [0.85 x %1.2f x %1.3f x (%1.6f x %3.1f)⅓ + 0.15 x %4.3f] x %4.1f x %4.1f" %(calc.Øc, calc.k, calc.ρs, calc.fck, calc.fn, calc.B, calc.D)
    wsout['C65'].value ="    = %10.0f N" %(calc.Vc) 
    wsout['D66'].value ="k = 1 + √(200 / d) = 1 + √(200 / %4.1f) = %1.3f (≤ 2.000)" %(calc.D, calc.k)
    wsout['D67'].value ="p = As / (b x d) = %6.1f / (%4.1f x %4.1f) = %1.6f ≤ 0.0200" %(calc.Asuse, calc.B, calc.D, calc.ρ) 
    wsout['D68'].value ="fn = Nu / Ac = %4.3f MPa (≤ 0.2 Φc fck = %4.3f MPa, 압축의경우+)" %(calc.fnn, calc.fnmax)
    wsout['C69'].value ="Vcd,min = (0.4 x Φc x fctk + 0.15 x fn] x b x d" 
    wsout['C70'].value ="        = (0.4 x %1.2f x %3.3f + 0.15 x %4.3f] x %4.1f x %4.1f" %(calc.Øc, calc.fctk, calc.fn, calc.B, calc.D)
    wsout['C71'].value ="        = %9.0f N" %(calc.Vcdmin)

    if calc.Vcd >= calc.Vun :
        wsout['C72'].value ="Vcd = %9.0f N ≥ Vu = %9.0f N  ∴ 전단보강 불필요" %(calc.Vcd, calc.Vun)
        wsout['B74'].value ="11) 최소 전단철근량 검토"
        wsout['C75'].value ="Av,use = %5.1f mm² ( %c %d - %2.3fea, C.T.C = %4.1f mm )" %(calc.Avs, calc.rebarid, calc.AvDia, calc.AvLeg, calc.AvSpace)
        wsout['C76'].value ="pv,min = (0.08 x √fck) / fvy = (0.08 x √%3.1f) / %4.1f = %1.6f" %(calc.fck, calc.fy, calc.ρvmin)
        if calc.ρvuse >= calc.ρvmin :
            wsout['C77'].value ="pv,use = %4.1f / (%4.1f x %4.1f x sin%3.1f) = %1.6f ≥ %1.6f ∴ O.K" %(calc.Avs, calc.AvSpace, calc.B, calc.αv, calc.ρvuse, calc.ρvmin)
        else :
            wsout['C77'].value ="pv,use = %4.1f / (%4.1f x %4.1f x sin%3.1f) = %1.6f < %1.6f ∴ N.G" %(calc.Avs, calc.AvSpace, calc.B, calc.αv, calc.ρvuse, calc.ρvmin)
        wsout['B79'].value ="12) 전단철근 간격검토"
        if calc.s1max >= calc.AvSpace :
            wsout['C80'].value ="- 종방향 Smax = 0.75 x %4.1f x (1 + cot%3.1f) = %4.1f ≥ %4.1f   ∴ O.K" %(calc.D, calc.αv, calc.s1max, calc.AvSpace)
        else :
            wsout['C80'].value ="- 종방향 Smax = 0.75 x %4.1f x (1 + cot%3.1f) = %4.1f < %4.1f   ∴ N.G" %(calc.D, calc.αv, calc.s1max, calc.AvSpace)    
        if calc.s2max >= calc.s2 :
            wsout['C81'].value ="- 횡방향 Smax = min(0.75 x %4.1f, 600.0) = %4.1f ≥ %4.1f   ∴ O.K" %(calc.D, calc.s2max, calc.s2)
        else :
            wsout['C81'].value ="- 종방향 Smax = min(0.75 x %4.1f, 600.0) = %4.1f < %4.1f   ∴ N.G" %(calc.D, calc.s2max, calc.s2)
    else :
        wsout['C72'].value ="Vcd = %9.0f N < Vu = %9.0f N  ∴ 전단보강 필요" %(calc.Vcd, calc.Vun)
        wsout['C74'].value ="cotΘ = %1.3f ( Θ = %3.1f˚)" %(calc.cotθ, calc.θ)
        if calc.cotθ < 1 :
            wsout['L74'].value ="1.0 > cotΘ → Θ 불만족  ∴ N.G"
        elif calc.cotθ > 2.5 :    
            wsout['L74'].value ="2.5 ≤ cotΘ → Θ 불만족  ∴ N.G"
        else :
            wsout['L74'].value ="1.0 ≤ cotΘ ≤ 2.5 → Θ 만족  ∴ O.K"
        wsout['C76'].value ="Av,use = %5.1f mm² ( %c %d - %2.3fea, C.T.C = %4.1f mm )" %(calc.Avs, calc.rebarid, calc.AvDia, calc.AvLeg, calc.AvSpace)
        wsout['C77'].value ="Vsd = (Φs x fvy x Av x z / s) x cotθ"
        wsout['C78'].value ="    = (%1.2f x %4.1f x %5.1f x %4.1f / %4.1f) x %1.3f" %(calc.Øs, calc.fy, calc.Avs, calc.z, calc.AvSpace, calc.cotθ)
        if calc.Vd >= calc.Vun :    
            wsout['C79'].value ="    = %d N ≥ Vu = %d N  ∴ O.K" %(calc.Vd, calc.Vun)
        else :
            wsout['C79'].value ="    = %d N < Vu = %d N  ∴ N.G" %(calc.Vd, calc.Vun)    
        wsout['C80'].value ="Vsd,max = (v x Φc x fck x b x z) / (cotθ + tanθ)"
        wsout['C81'].value ="        = (%1.3f x %1.2f x %3.1f x %4.1f x %4.1f) / ( %1.3f + %1.3f )" %(calc.ν, calc.Øc, calc.fck, calc.B, calc.z, calc.cotθ, calc.tanθ)
        if calc.Vd <= calc.Vdmax :
            wsout['C82'].value ="        = %d N ≤ Vsd = %d N  ∴ O.K" %(calc.Vdmax, calc.Vd)
        else :
            wsout['C82'].value ="        = %d N > Vsd = %d N  ∴ N.G" %(calc.Vdmax, calc.Vd)
        wsout['B84'].value ="11) 최소 전단철근량 검토"
        wsout['C85'].value ="pv,min = (0.08 x √fck) / fvy = (0.08 x √%3.1f) / %4.1f = %f" %(calc.fck, calc.fy, calc.ρvmin)
        if calc.ρvuse >= calc.ρvmin :
            wsout['C86'].value ="pv,use = %4.1f / (%4.1f x %4.1f x sin%3.1f) = %1.6f ≥ %1.6f ∴ O.K" %(calc.Avs, calc.AvSpace, calc.B, calc.αv, calc.ρvuse, calc.ρvmin)
        else :
            wsout['C86'].value ="pv,use = %4.1f / (%4.1f x %4.1f x sin%3.1f) = %1.6f < %1.6f ∴ N.G" %(calc.Avs, calc.AvSpace, calc.B, calc.αv, calc.ρvuse, calc.ρvmin)
        wsout['B88'].value ="12) 전단철근 간격검토"
        if calc.s1max >= calc.AvSpace :
            wsout['C89'].value ="- 종방향 Smax = 0.75 x %4.1f x (1 + cot%3.1f) = %4.1f ≥ %4.1f   ∴ O.K" %(calc.D, calc.αv, calc.s1max, calc.AvSpace)
        else :
            wsout['C89'].value ="- 종방향 Smax = 0.75 x %4.1f x (1 + cot%3.1f) = %4.1f < %4.1f   ∴ N.G" %(calc.D, calc.αv, calc.s1max, calc.AvSpace)    
        if calc.s2max >= calc.s2 :
            wsout['C90'].value ="- 횡방향 Smax = min(0.75 x %4.1f, 600.0) = %4.1f ≥ %4.1f   ∴ O.K" %(calc.D, calc.s2max, calc.s2)
        else :
            wsout['C90'].value ="- 횡방향 Smax = min(0.75 x %4.1f, 600.0) = %4.1f < %4.1f   ∴ N.G" %(calc.D, calc.s2max, calc.s2) 
        wsout['B92'].value ="13) 작용전단력에 의해 종방향 철근에 발생하는 추가인장력 검토"
        wsout['C93'].value ="ΔTr = (Mr - Mu) / z = (%10.1f - %10.1f) / %4.1f = %9.1f N" %(calc.Mr, calc.Mun, calc.D, calc.dTr)
        wsout['C94'].value ="ΔT = 0.5 x Vu x (cotθ - cotα) = 0.5 x %9.1f x (cot%3.1f - cot%3.1f)" %(calc.Vun, calc.θ, calc.αv)
        if calc.dT >= calc.dTr :
            wsout['C95'].value ="     = %9.1f N ≥ %9.1f N  → ∴ 추가 주철근 필요" %(calc.dT, calc.dTr)
        else :
            wsout['C95'].value ="     = %9.1f N < %9.1f N  → ∴ 추가 주철근 필요 없음" %(calc.dT, calc.dTr)

sexcelsheet(calc, wsout)
sexcelsheet(calc1, wsout1)
#--------------------------------------
#          사용성 검토(균열검토)
#--------------------------------------
calc.calservice()
def vexcelsheet(calc, wsout) :
    wsout['B100'].value ="14) 균열검토"
    wsout['B101'].value =" - 한계상태 검증을 위한 설계등급 : E (표면 한계균열폭 0.3 mm)"
    wsout['B102'].value =" ① 비균열 가정시 인장 연단 응력(사용하중조합-Ⅰ)"
    wsout['C103'].value ="ft = M / Zb = M / (b x h² / 6))"
    wsout['C104'].value ="   = %10.1f / (%4.1f x %4.1f² / 6))" %(calc.Msn1, calc.B, calc.H)
    if calc.fct1 < calc.fctk :
        wsout['C105'].value ="   = %10.1f mm² < %4.1f mm²  → ∴ 비균열단면으로 균열검토 필요없음" %(calc.fct1, calc.fctk)
    else :
        wsout['C105'].value ="   = %10.1f mm² ≥ %4.1f mm²  → ∴ 균열단면으로 균열검토 필요" %(calc.fct1, calc.fctk)
        wsout['B106'].value =" ② 균열응력 검토"
        wsout['C107'].value ="n = Es / Ec = %d" %(calc.nr)
        wsout['C108'].value ="c = -n x As / b + n x As / b x √(1 + 2 x b x d / (n x As))"
        wsout['C109'].value ="  = -%d x %4.1f / %4.1f + %d x %6.1f / %4.1f x √(1 + 2 x %4.1f x %4.1f / (%d x %4.1f))" %(calc.nr, calc.Asuse, calc.B, calc.nr, calc.Asuse, calc.B, calc.B, calc.D, calc.nr, calc.Asuse)
        wsout['C110'].value ="  = %2.3f mm" %(calc.c)
        wsout['C111'].value ="- 콘크리트 연단에 발생하는 응력" 
        wsout['D112'].value ="fc1 = 2 x M1 / ((1 - k / 3) x b x c x d)"
        wsout['D113'].value ="    = 2 x %10.0f / ((1 - %4.1f / 3) x %4.1f x %2.3f x %4.1f)" %(calc.Msn1, calc.k, calc.B, calc.c, calc.D)
        if calc.fc1 < calc.fca1 :
            wsout['D114'].value ="    = %4.3f N/mm² < 0.6 x fck = %4.1f N/mm²   ∴ O.k" %(calc.fc1, calc.fca1)
        else :
            wsout['D114'].value ="    = %4.3f N/mm² ≥ 0.6 x fck = %4.1f N/mm²   ∴ N.G" %(calc.fc1, calc.fca1)
        wsout['D115'].value ="fc5 = 2 x M5 / ((1 - k / 3) x b x c x d)"
        wsout['D116'].value ="    = 2 x %10.0f / ((1 - %4.1f / 3) x %4.1f x %2.3f x %4.1f)" %(calc.Msn5, calc.k, calc.B, calc.c, calc.D)
        if calc.fc1 < calc.fca1 :
            wsout['D117'].value ="    = %4.3f N/mm² < 0.45 x fck = %4.1f N/mm²   ∴ O.k" %(calc.fc5, calc.fca5)
        else :
            wsout['D117'].value ="    = %4.3f N/mm² ≥ 0.45 x fck = %4.1f N/mm²   ∴ N.G" %(calc.fc5, calc.fca5)
        wsout['C118'].value ="- 철근에 발생하는 응력" 
        wsout['D119'].value =" fs1 = M1 / (As x (1 - k / 3) x d)"
        wsout['D120'].value ="     = %10.1f / (%6.1f x (1 - %4.1f / 3) x %4.1f)" %(calc.Msn1, calc.Asuse, calc.k, calc.D)
        if calc.fs1 < calc.fsa :
            wsout['D121'].value ="     = %4.3f N/mm² < %4.1f N/mm²   ∴ O.k" %(calc.fs1, calc.fsa)
        else :
            wsout['D121'].value ="     = %4.3f N/mm² ≥ %4.1f N/mm²   ∴ N.G" %(calc.fs1, calc.fsa)
        wsout['B123'].value =" ③ 균열제어를 위한 최소철근량"
        wsout['C124'].value ="As,min = kc x k x Act x (fct / fs) [도.설(2015) 5.8.3.2]" 
        wsout['C125'].value ="       = %2.3f x %2.3f x %6.1f x (%2.3f / %4.1f) = %6.1fmm²" %(calc.kc, calc.k, calc.Act, calc.fct, calc.fsa,calc.Asdmin) 
        wsout['C126'].value =" - Act = min[2.5 x (h - d), (h - x)/3, h/2] x B"
        wsout['C127'].value ="       = min[2.5 x (%4.1f - %4.1f), (%4.1f - %4.1f)/3, %4.1f/2] x %4.1f" %(calc.H, calc.D, calc.H, calc.x, calc.H, calc.B)
        wsout['C128'].value ="       = %6.1f mm²" %(calc.Act)
        wsout['C129'].value =" -  x = - n x As / b + n x As / b x √(1 + 2 x b x d / (n x As))"
        wsout['C130'].value ="      = - %d x %6.1f / %4.1f + %d x %6.1f / %4.1f x √(1 + 2 x %4.1f x %4.1f / (%d x %6.1f))" %(calc.nr, calc.Asuse, calc.B, calc.nr, calc.Asuse, calc.B, calc.B, calc.D, calc.nr, calc.Asuse)
        wsout['C131'].value ="      = %3.2f mm" %(calc.x)
        if calc.H <= 300:    
            wsout['C132'].value =" -  k = %2.3f ( H ≤ 300 )" %(calc.k)
        elif calc.H < 800:
            wsout['C132'].value =" -  k = -0.0007 x %3.1f + 1.21 = %2.3f ( H < 800 )" %(calc.H, calc.k)
        else :
            wsout['C132'].value =" -  k = %2.3f ( H ≥ 800 )" %(calc.k)
        wsout['C133'].value =" -  kc = 0.4 x [1 - fn / (k1 x (h / h') x fct)] ≤ 1"
        wsout['C134'].value ="       = 0.4 x [1 - %3.3f / (%1.3f x (%4.1f / %4.1f) x %2.3f)]" %(calc.fnn, calc.k1, calc.H, calc.h1, calc.fct)
        if calc.Asuse >= calc.Asdmin :
            wsout['C135'].value ="As,use = %6.1f mm² ≥ As,min = %6.1f mm²    ∴ O.K" %(calc.Asuse, calc.Asdmin)
        else :
            wsout['C135'].value ="As,use = %6.1f mm² < As,min = %6.1f mm²    ∴ N.G" %(calc.Asuse, calc.Asdmin)
        wsout['B137'].value =" ④ 간접균열검토"
        wsout['C138'].value =" - fs5 = Ms5 / (As x (1 - k / 3) x d)"
        wsout['C139'].value ="       = %10.1f / (%6.1f x (1 - %4.1f / 3) x %4.1f) = %4.3f N/mm²" %(calc.Msn5, calc.Asuse, calc.k, calc.D, calc.fs5)
        wsout['C140'].value =" - 사용철근 직경 : D = %dmm, fs = %4.1fN/mm²" %(calc.AsDia1, calc.fsad)
        wsout['C141'].value =" - 사용철근 간격 : s = %dmm, fs = %4.1fN/mm²" %(calc.Asspace1, calc.fsas)
        if calc.fsaf >= calc.fs5 :
            wsout['C142'].value =" - 철근 한계응력 : MAX(%4.1f, %4.1f) ≥ fs = %4.1f N/mm²  ∴ O.K" %(calc.fsad, calc.fsas, calc.fs5)
        else :
            wsout['C142'].value =" - 철근 한계응력 : MAX(%4.1f, %4.1f) < fs = %4.1f N/mm²  ∴ N.G" %(calc.fsad, calc.fsas, calc.fs5)

vexcelsheet(calc, wsout)   
wfname = filedialog.asksaveasfile(title="Save result",filetypes=[("Excel files", "*.xlsx")]).name
wb.save(wfname)
#wb.save('Calc_As_Output_3.0.xlsx')
wb.close()
